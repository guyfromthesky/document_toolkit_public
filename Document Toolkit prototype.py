#System variable and io handling
import os
from posixpath import basename
import sys

import configparser
#Regular expression handlings
import multiprocessing
from multiprocessing import Process , Queue, Manager
import queue 
import subprocess
#Get timestamp
import time
from datetime import datetime
#function difination
import unicodedata
from urllib.parse import urlparse

#GUI
from tkinter.ttk import Entry, Label, Notebook, Progressbar, Frame
from tkinter.ttk import Checkbutton, Button, Radiobutton

from tkinter import Tk, Frame
from tkinter import Menu, filedialog, messagebox
from tkinter import Text, colorchooser
from tkinter import IntVar, StringVar
from tkinter import W, E, S, N, END, HORIZONTAL
from tkinter import WORD, NORMAL, ACTIVE, INSERT
from tkinter import DISABLED

import webbrowser

import pandas as pd
import numpy as np

from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font
from openpyxl.styles import PatternFill
from openpyxl.styles import Color
from openpyxl.styles import Color, PatternFill, Font

from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter, column_index_from_string


#from document_toolkit_function.py import *

DELAY1 = 20

ToolDisplayName = "Document Toolkit"
tool_name = 'document'
rev = 1302
a,b,c,d = list(str(rev))
VerNum = a + '.' + b + '.' + c + chr(int(d)+97)

version = ToolDisplayName  + " " +  VerNum 

#**********************************************************************************
# UI handle ***********************************************************************
#**********************************************************************************

class Document_Utility(Frame):
	def __init__(self, Root, Queue = None, Manager = None,):
		
		Frame.__init__(self, Root) 
		#super().__init__()
		self.parent = Root 

		# Queue
		self.Process_Queue = Queue['Process_Queue']
		self.Result_Queue = Queue['Result_Queue']
		self.Status_Queue = Queue['Status_Queue']
		self.Debug_Queue = Queue['Debug_Queue']

		self.Manager = Manager['Default_Manager']

		self.Config_Init()

		self.Options = {}

		# XLSX Optmizer
		self.Optimize_Folder = ""
		self.Optimize_FileList = ""
		# XLSX Comparision
		self.Compare_Folder_Old = ""
		self.Compare_File_List_Old = ""
		self.Compare_Folder_New = ""
		self.Compare_File_List_New = ""

		# UI Variable
		self.Button_Width_Full = 20
		self.Button_Width_Half = 15
		
		self.PadX_Half = 5
		self.PadX_Full = 10
		self.PadY_Half = 5
		self.PadY_Full = 10
		self.StatusLength = 120
		self.AppLanguage = 'en'

		
		self.App_LanguagePack = {}
		self.initGeneralSetting()

		if self.AppLanguage != 'kr':
			from languagepack import LanguagePackEN as LanguagePack
		else:
			from languagepack import LanguagePackKR as LanguagePack

		self.LanguagePack = LanguagePack

		# Init function

		self.parent.resizable(False, False)
		self.parent.title(version)
		# Creating Menubar 
		
		#**************New row#**************#
		self.Notice = StringVar()
		self.Debug = StringVar()
		self.Progress = StringVar()
	
		
		self.basePath = os.path.abspath(os.path.dirname(sys.argv[0]))
		self.ExceptionPath = self.basePath + "\\Exception.xlsx"
		try:
			self.ExceptionList = self.ImportException(self.ExceptionPath)
			print('My exception list: ', self.ExceptionList)
		except:
			self.ExceptionList = []
		#Generate UI

		self.Generate_Menu_UI()
		self.Generate_Tab_UI()
		self.init_UI()
		self.init_UI_Configuration()

	def Config_Init(self):
		self.Roaming = os.environ['APPDATA'] + '\\Document_Utility'
		self.AppConfig = self.Roaming + '\\config.ini'
	
		if not os.path.isdir(self.Roaming):
			try:
				os.mkdir(self.Roaming)
			except OSError:
				print ("Creation of the directory %s failed" % self.Roaming)
		else:
			print('Roaming folder exist.')



	# UI init
	def init_UI(self):
	
		self.Generate_Auto_Test_UI(self.AutoTest)
		
		self.Generate_Fast_Comparision_UI(self.FastComparison)
		
		self.Generate_Data_Deep_Comparision_UI(self.DeepDataCompare)

		self.Generate_Optimizer_UI(self.Optimizer)

		self.Generate_Profanity_Detector_UI(self.ProfanityDetector)

		#self.Generate_Language_Validator_UI(self.LanguageValidator)
		
		self.Generate_Debugger_UI(self.Process)
		
		# Debugger

	def Generate_Menu_UI(self):
		menubar = Menu(self.parent) 
		# Adding File Menu and commands 
		'''
		file = Menu(menubar, tearoff = 0)
		
		# Adding Load Menu 
		menubar.add_cascade(label =  self.LanguagePack.Menu['File'], menu = file) 
		file.add_command(label =  self.LanguagePack.Menu['LoadTM'], command = self.Menu_Function_Select_TM) 
		file.add_separator() 
		file.add_command(label =  self.LanguagePack.Menu['CreateTM'], command = self.Menu_Function_Create_TM)
		file.add_separator() 
		file.add_command(label =  self.LanguagePack.Menu['Exit'], command = self.parent.destroy) 
		'''
		# Adding Help Menu
		help_ = Menu(menubar, tearoff = 0) 
		menubar.add_cascade(label =  self.LanguagePack.Menu['Help'], menu = help_) 
		help_.add_command(label =  self.LanguagePack.Menu['GuideLine'], command = self.Menu_Function_Open_Main_Guideline) 
		help_.add_separator()
		help_.add_command(label =  self.LanguagePack.Menu['About'], command = self.Menu_Function_About) 
		self.parent.config(menu = menubar)

		# Adding Help Menu
		language = Menu(menubar, tearoff = 0) 
		menubar.add_cascade(label =  self.LanguagePack.Menu['Language'], menu = language) 
		language.add_command(label =  self.LanguagePack.Menu['Hangul'], command = self.SetLanguageKorean) 
		language.add_command(label =  self.LanguagePack.Menu['English'], command = self.SetLanguageEnglish) 
		self.parent.config(menu = menubar) 	

	def Generate_Tab_UI(self):
		self.TAB_CONTROL = Notebook(self.parent)
		#Tab
		
		#self.MultiDataCompare = ttk.Frame(self.TAB_CONTROL)
		#self.TAB_CONTROL.add(self.MultiDataCompare, text= self.LanguagePack.Tab['MultiDeepCompare'])

		self.AutoTest = Frame(self.TAB_CONTROL)
		self.TAB_CONTROL.add(self.AutoTest, text= self.LanguagePack.Tab['AutomationTest'])

		self.FastComparison = Frame(self.TAB_CONTROL)
		self.TAB_CONTROL.add(self.FastComparison, text= self.LanguagePack.Tab['FastCompare'])

		self.DeepDataCompare = Frame(self.TAB_CONTROL)
		self.TAB_CONTROL.add(self.DeepDataCompare, text= self.LanguagePack.Tab['StructuredCompare'])
		
		#Tab
		self.Optimizer = Frame(self.TAB_CONTROL)
		self.TAB_CONTROL.add(self.Optimizer, text= self.LanguagePack.Tab['Optimizer'])
		
		#Tab
		'''
		self.LanguageValidator = Frame(self.TAB_CONTROL)
		self.TAB_CONTROL.add(self.LanguageValidator, text= 'Language Validator')
		'''
		

		self.ProfanityDetector = Frame(self.TAB_CONTROL)
		self.TAB_CONTROL.add(self.ProfanityDetector, text= self.LanguagePack.Tab['BadWordTest'])
		
		#Tab
		'''
		self.DataLookup = ttk.Frame(TAB_CONTROL)
		TAB_CONTROL.add(self.DataLookup, text=  'Data Lookup')
		'''	
		#Tab
		self.Process = Frame(self.TAB_CONTROL)
		self.TAB_CONTROL.add(self.Process, text= self.LanguagePack.Tab['Debug'])
		
		self.TAB_CONTROL.pack(expand=1, fill="both")
		return

	#STABLE
	def Generate_Data_Deep_Comparision_UI(self, Tab):
		
		Row = 1
		Label(Tab, textvariable=self.Notice).grid(row=Row, column=1, columnspan = 10, padx=5, pady=5, sticky= W)
		Row += 1

		self.Str_Deep_Old_File_Path = StringVar()
		Label(Tab, text=  self.LanguagePack.Label['OldDataTable']).grid(row=Row, column=1, columnspan=2, padx=5, pady=5, sticky= W)
		self.Entry_Old_File_Path = Entry(Tab,width = 110, state="readonly", textvariable=self.Str_Deep_Old_File_Path)
		self.Entry_Old_File_Path.grid(row=Row, column=3, columnspan=4, padx=4, pady=5, sticky=E)
		Button(Tab, width = self.Button_Width_Half, text=  self.LanguagePack.Button['Browse'], command= self.Btn_Browse_Old_Data_File).grid(row=Row, column=7, columnspan=2, padx=5, pady=5, sticky=E)
		Button(Tab, width = self.Button_Width_Half, text=  self.LanguagePack.Button['SelectBGColor'], command= self.Btn_Select_Background_Colour).grid(row=Row, column=9, columnspan=2,padx=5, pady=5, sticky=W)

		Row += 1
		self.Str_Deep_New_File_Path = StringVar()
		Label(Tab, text=  self.LanguagePack.Label['NewDataTable']).grid(row=Row, column=1, columnspan=2, padx=5, pady=5, sticky= W)
		self.Entry_New_File_Path = Entry(Tab,width = 110, state="readonly", textvariable=self.Str_Deep_New_File_Path)
		self.Entry_New_File_Path.grid(row=Row, column=3, columnspan=4, padx=4, pady=5, sticky=E)
		Button(Tab, width = self.Button_Width_Half, text=  self.LanguagePack.Button['Browse'], command= self.Btn_Browse_New_Data_File).grid(row=Row, column=7, columnspan=2, padx=5, pady=5, sticky=E)
		Button(Tab, width = self.Button_Width_Half, text=  self.LanguagePack.Button['SelectFontColor'], command= self.Btn_Select_Font_Colour).grid(row=Row, column=9, columnspan=2,padx=5, pady=5, sticky=W)
		
	
		Row += 1
		Label(Tab, text=self.LanguagePack.Label['Main_Data_Sheet']).grid(row=Row, column=1, padx=5, pady=5, sticky=W)
		self.Str_Data_Sheet_Name = Text(Tab, width = 110, height=1) #
		self.Str_Data_Sheet_Name.grid(row=Row, column=2, columnspan=8, padx=5, pady=5, sticky=E)
		self.Str_Data_Sheet_Name.insert("end", 'Data')
		
		Row += 1
		Label(Tab, text=self.LanguagePack.Label['ID_Col']).grid(row=Row, column=1, padx=5, pady=5, sticky=W)
		self.Str_Data_Col_Name = Text(Tab, width = 110, height=1) #
		self.Str_Data_Col_Name.grid(row=Row, column=2, columnspan=8, padx=5, pady=5, sticky=E)
		self.Str_Data_Col_Name.insert("end", 'StringId')
		
		Row+=1
		Label(Tab, text=  None).grid(row=Row, column=1)	
		Row+=1
		Label(Tab, text=  None).grid(row=Row, column=1)	
		Row+=1
		Label(Tab, text=  None).grid(row=Row, column=1)	
		Row += 1
		Button(Tab, width = self.Button_Width_Half, text=  self.LanguagePack.Button['OpenOutput'], command= self.Open_Deep_Compare_Result_Folder).grid(row=Row, column=7, columnspan=2,padx=5, pady=5, sticky=W)
		Button(Tab, width = self.Button_Width_Half, text=  self.LanguagePack.Button['Execute'], command= self.Btn_Deep_Compare_Data).grid(row=Row, column=9, columnspan=2,padx=5, pady=5, sticky=W)

	def Generate_Auto_Test_UI(self, Tab):
		
		Row = 1
		Label(Tab, textvariable=self.Notice).grid(row=Row, column=1, columnspan = 10, padx=5, pady=5, sticky= W)
		Row += 1

		self.Str_AutoTest_Old_Folder_Path = StringVar()
		self.AutoTest_Old_Folder_Path = None
		Label(Tab, text=  self.LanguagePack.Label['OldDataTable']).grid(row=Row, column=1, columnspan=2, padx=5, pady=5, sticky= W)
		self.Entry_Old_File_Path = Entry(Tab,width = 110, state="readonly", textvariable=self.Str_AutoTest_Old_Folder_Path)
		self.Entry_Old_File_Path.grid(row=Row, column=3, columnspan=4, padx=4, pady=5, sticky=E)
		Button(Tab, width = self.Button_Width_Half, text=  self.LanguagePack.Button['Browse'], command= self.Btn_AutoTest_Browse_Old_Folder).grid(row=Row, column=7, columnspan=2, padx=5, pady=5, sticky=E)
		Button(Tab, width = self.Button_Width_Half, text=  self.LanguagePack.Button['SelectBGColor'], command= self.Btn_AutoTest_Select_Background_Colour).grid(row=Row, column=9, columnspan=2,padx=5, pady=5, sticky=W)

		Row += 1
		self.Str_AutoTest_New_Folder_Path = StringVar()
		self.AutoTest_New_Folder_Path = None
		Label(Tab, text=  self.LanguagePack.Label['NewDataTable']).grid(row=Row, column=1, columnspan=2, padx=5, pady=5, sticky= W)
		self.Entry_New_File_Path = Entry(Tab,width = 110, state="readonly", textvariable=self.Str_AutoTest_New_Folder_Path)
		self.Entry_New_File_Path.grid(row=Row, column=3, columnspan=4, padx=4, pady=5, sticky=E)
		Button(Tab, width = self.Button_Width_Half, text=  self.LanguagePack.Button['Browse'], command= self.Btn_AutoTest_Browse_New_Folder).grid(row=Row, column=7, columnspan=2, padx=5, pady=5, sticky=E)
		Button(Tab, width = self.Button_Width_Half, text=  self.LanguagePack.Button['SelectFontColor'], command= self.Btn_AutoTest_Select_Font_Colour).grid(row=Row, column=9, columnspan=2,padx=5, pady=5, sticky=W)
		

		Row += 1
		self.Str_Test_Configuration_File_Path = StringVar()
		Label(Tab, text=  self.LanguagePack.Label['AutoTestConfig']).grid(row=Row, column=1, columnspan=2, padx=5, pady=5, sticky= W)
		temp = os.getcwd()
		
		self.List_File_Test = Entry(Tab,width = 110, state="readonly", textvariable=self.Str_Test_Configuration_File_Path)
		self.List_File_Test.grid(row=Row, column=3, columnspan=4, padx=5, pady=5, sticky=E)

		self.test_configuration_file = ''
		Button(Tab, width = self.Button_Width_Half, text=  self.LanguagePack.Button['Browse'], command= self.Btn_AutoTest_Browse_Test_Config).grid(row=Row, column=7, columnspan=2, padx=5, pady=5, sticky=E)
		Button(Tab, width = self.Button_Width_Half, text=  self.LanguagePack.Button['Delete'], command= self.Btn_AutoTest_Clear_Selected_File).grid(row=Row, column=9, columnspan=2, padx=5, pady=5, sticky=E)

		Row+=1

		Label(Tab, text=  self.LanguagePack.Label['Progress']).grid(row=Row, column=1, columnspan=2, padx=5, pady=5, sticky= W)

		self.progressbar = Progressbar(Tab, orient=HORIZONTAL, length=900,  mode='determinate')
		self.progressbar["maximum"] = 1000
		self.progressbar.grid(row=Row, column=3, columnspan=7, padx=5, pady=5, sticky=E+W)

		Row+=1
		Label(Tab, text=  None).grid(row=Row, column=1)
		Row+=1
		Label(Tab, text=  None).grid(row=Row, column=1)
		Row+=1
		Label(Tab, text=  None).grid(row=Row, column=1)
		Row+=1
		Label(Tab, text=  None).grid(row=Row, column=1)
		Row+=1
		Label(Tab, text=  None).grid(row=Row, column=1)
		Row += 1
		Button(Tab, width = self.Button_Width_Half, text=  self.LanguagePack.Button['OpenOutput'], command= self.Btn_Open_AutoTest_Result_Folder).grid(row=Row, column=7, columnspan=2,padx=5, pady=5, sticky=W)
		Button(Tab, width = self.Button_Width_Half, text=  self.LanguagePack.Button['Execute'], command= self.Btn_AutoTest_Execute).grid(row=Row, column=9, columnspan=2,padx=5, pady=5, sticky=W)


	#NOT STABLE
	def Generate_Fast_Comparision_UI(self, Tab):
		
		Row = 1
		Label(Tab, textvariable=self.Notice).grid(row=Row, column=1, columnspan = 10, padx=5, pady=5, sticky= W)
		Row += 1

		self.Str_Old_File_Path = StringVar()
		Label(Tab, text=  self.LanguagePack.Label['OldDataTable']).grid(row=Row, column=1, columnspan=2, padx=5, pady=5, sticky= W)
		self.Entry_Old_File_Path = Entry(Tab,width = 110, state="readonly", textvariable=self.Str_Old_File_Path)
		self.Entry_Old_File_Path.grid(row=Row, column=3, columnspan=4, padx=4, pady=5, sticky=E)
		Button(Tab, width = self.Button_Width_Half, text=  self.LanguagePack.Button['Browse'], command= self.Btn_FastCompare_Browse_Old_Data_Folder).grid(row=Row, column=7, columnspan=2, padx=5, pady=5, sticky=E)
		Button(Tab, width = self.Button_Width_Half, text=  self.LanguagePack.Button['SelectBGColor'], command= self.Btn_Select_Background_Colour).grid(row=Row, column=9, columnspan=2,padx=5, pady=5, sticky=W)

		Row += 1
		self.Str_New_File_Path = StringVar()
		Label(Tab, text=  self.LanguagePack.Label['NewDataTable']).grid(row=Row, column=1, columnspan=2, padx=5, pady=5, sticky= W)
		self.Entry_New_File_Path = Entry(Tab,width = 110, state="readonly", textvariable=self.Str_New_File_Path)
		self.Entry_New_File_Path.grid(row=Row, column=3, columnspan=4, padx=4, pady=5, sticky=E)
		Button(Tab, width = self.Button_Width_Half, text=  self.LanguagePack.Button['Browse'], command= self.Btn_FastCompare_Browse_New_Data_Folder).grid(row=Row, column=7, columnspan=2, padx=5, pady=5, sticky=E)
		Button(Tab, width = self.Button_Width_Half, text=  self.LanguagePack.Button['SelectFontColor'], command= self.Btn_Select_Font_Colour).grid(row=Row, column=9, columnspan=2,padx=5, pady=5, sticky=W)
		
	
		Row += 1
		Label(Tab, text=self.LanguagePack.Label['Main_Data_Sheet']).grid(row=Row, column=1, padx=5, pady=5, sticky=W)
		self.Str_FastCompare_Data_Sheet_Name = Text(Tab, width = 110, height=1) #
		self.Str_FastCompare_Data_Sheet_Name.grid(row=Row, column=2, columnspan=8, padx=5, pady=5, sticky=E)
		self.Str_FastCompare_Data_Sheet_Name.insert("end", 'Data')
		
		Row += 1
		Label(Tab, text=self.LanguagePack.Label['ID_Col']).grid(row=Row, column=1, padx=5, pady=5, sticky=W)
		self.Str_FastCompare_Data_Col_Name = Text(Tab, width = 110, height=1) #
		self.Str_FastCompare_Data_Col_Name.grid(row=Row, column=2, columnspan=8, padx=5, pady=5, sticky=E)
		self.Str_FastCompare_Data_Col_Name.insert("end", 'StringId')
		
		Row += 1
		self.Str_List_File_Test = StringVar()
		Label(Tab, text=  self.LanguagePack.Label['ListFile']).grid(row=Row, column=1, columnspan=2, padx=5, pady=5, sticky= W)
		self.List_File_Test = Entry(Tab,width = 110, state="readonly", textvariable=self.Str_List_File_Test)
		self.List_File_Test.grid(row=Row, column=3, columnspan=4, padx=4, pady=5, sticky=E)
		#self.List_Test_Files
		self.List_Test_Files = ''
		Button(Tab, width = self.Button_Width_Half, text=  self.LanguagePack.Button['Browse'], command= self.Btn_FastCompare_Browse_List_File).grid(row=Row, column=7, columnspan=2, padx=5, pady=5, sticky=E)
		Button(Tab, width = self.Button_Width_Half, text=  self.LanguagePack.Button['Delete'], command= self.Btn_FastCompare_Clear_Selected_File).grid(row=Row, column=9, columnspan=2, padx=5, pady=5, sticky=E)

		Row+=1
		self.Skip_Comment_Column = IntVar()
		SkipCommentButton = Checkbutton(Tab, text= 'Skip # Col', variable = self.Skip_Comment_Column, command= None)
		SkipCommentButton.grid(row=Row, column=1, columnspan=2, padx=5, pady=5, sticky=W)
		SkipCommentButton.bind("<Enter>", lambda event : self.Notice.set(self.LanguagePack.ToolTips['TMTranslate']))
		self.Skip_Comment_Column.set(1)

		Row+=1
		Label(Tab, text=  None).grid(row=Row, column=1)	
		Row += 1
		Button(Tab, width = self.Button_Width_Half, text=  self.LanguagePack.Button['OpenOutput'], command= self.Open_Fast_Compare_Result_Folder).grid(row=Row, column=7, columnspan=2,padx=5, pady=5, sticky=W)
		Button(Tab, width = self.Button_Width_Half, text=  self.LanguagePack.Button['Execute'], command= self.Btn_Fast_Compare_Data).grid(row=Row, column=9, columnspan=2,padx=5, pady=5, sticky=W)


	#STABLE
	def Generate_Optimizer_UI(self, Tab):
		
		Row = 1
		Label(Tab, textvariable=self.Notice).grid(row=Row, column=1, columnspan = 10, padx=5, pady=5, sticky= W)
		Row += 1
		self.RawSource = StringVar()
		Label(Tab, text=  self.LanguagePack.Label['OptimizeDatafile'],  width = self.Button_Width_Half).grid(row=Row, column=1, columnspan=2, padx=5, pady=5, sticky= W)
	
		self.TextRawSourcePath = Entry(Tab,width = 130, state="readonly", textvariable=self.RawSource)
		self.TextRawSourcePath.grid(row=Row, column=3, columnspan=6, padx=5, pady=5, sticky=W)
		Button(Tab, width = self.Button_Width_Half, text=  self.LanguagePack.Button['Browse'], command= self.Btn_Load_Raw_Source).grid(row=Row, column=9, columnspan=2, padx=5, pady=5, sticky=E)
		Row+=1
		Button(Tab, width = self.Button_Width_Half, text=  self.LanguagePack.Button['Execute'], command= self.Btn_Optimize_XLSX).grid(row=Row, column=9, columnspan=2,padx=5, pady=5, sticky=W)
		Row+=1
		Button(Tab, width = self.Button_Width_Half, text=  self.LanguagePack.Button['OpenOutput'], command= self.Btn_Open_Optimizer_Folder).grid(row=Row, column=9, columnspan=2,padx=5, pady=5, sticky=W)

		Row+=1
		Label(Tab, text=  None).grid(row=Row, column=1)	
		Row+=1
		Label(Tab, text=  None).grid(row=Row, column=1)	
		Row+=1
		Label(Tab, text=  None).grid(row=Row, column=1)	
		Row+=1
		Label(Tab, text=  None).grid(row=Row, column=1)	

		
		self.Optimize_Progressbar = Progressbar(Tab, orient=HORIZONTAL, length=1000,  mode='determinate')
		self.Optimize_Progressbar["maximum"] = 1000
		self.Optimize_Progressbar.grid(row=10, column=1, columnspan=9, padx=5, pady=5, sticky=W)	

	
	def Generate_Profanity_Detector_UI(self, Tab):
		
		Row = 1
		Label(Tab, textvariable=self.Notice).grid(row=Row, column=1, columnspan = 10, padx=5, pady=5, sticky= W)
		Row += 1

		self.Str_Text_File_Path = StringVar()
		Label(Tab, text= self.LanguagePack.Label['TextFile']).grid(row=Row, column=1, columnspan=2, padx=5, pady=5, sticky= W)

		self.BadWord_Text_Source = Entry(Tab, width = 120, state="readonly", textvariable=self.Str_Text_File_Path)
		self.BadWord_Text_Source.grid(row=Row, column=3, columnspan=4, padx=4, pady=5, sticky=W)
		
		#Browse data (text) files
		Button(Tab, width = self.Button_Width_Full, text=  self.LanguagePack.Button['Browse'], command= self.Btn_BadWord_Browse_Data_Files).grid(row=Row, column=9, columnspan=2, padx=5, pady=5, sticky=E)
		
		Row += 1

		self.Str_BadWord_DB_Path = StringVar()
		Label(Tab, text= self.LanguagePack.Label['DBTextFile']).grid(row=Row, column=1, columnspan=2, padx=5, pady=5, sticky= W)

		self.BadWord_DB_Source = Entry(Tab, width = 120, state="readonly", textvariable=self.Str_BadWord_DB_Path)
		self.BadWord_DB_Source.grid(row=Row, column=3, columnspan=4, padx=4, pady=5, sticky=W)
		
		#Browse DB file
		Button(Tab, width = self.Button_Width_Full, text=  self.LanguagePack.Button['Browse'], command= self.Btn_BadWord_Browse_DB_File).grid(row=Row, column=9, columnspan=2, padx=5, pady=5, sticky=E)

		Row += 1
		Label(Tab, text=self.LanguagePack.Label['Main_Data_Sheet']).grid(row=Row, column=1, padx=5, pady=5, sticky=W)
		self.BadWord_Data_Sheet_Name = Text(Tab, width = 90, height=1) #
		self.BadWord_Data_Sheet_Name.grid(row=Row, column=3, columnspan=4, padx=5, pady=5, sticky=W)
		self.BadWord_Data_Sheet_Name.insert("end", 'Data')
		Button(Tab, width = self.Button_Width_Full, text=  self.LanguagePack.Button['OpenOutput'], command= self.Open_BadWord_Result_Folder).grid(row=Row, column=9, columnspan=2,padx=5, pady=5, sticky=W)


		Row += 1
		Label(Tab, text= self.LanguagePack.Label['TextColumn']).grid(row=Row, column=1, padx=5, pady=5, sticky=W)
		self.BadWord_ColumnID = Text(Tab, width = 90, height=1) #
		self.BadWord_ColumnID.grid(row=Row, column=3, columnspan=4, padx=5, pady=5, sticky=W)
		self.BadWord_ColumnID.insert("end", 'String')
		

		'''
		Row += 1
		Label(Tab, text="Search Key: ").grid(row=Row, column=1, padx=5, pady=5, sticky=W)
		self.ShowColumnID = Text(Tab, width = 90, height=1) #
		self.ShowColumnID.grid(row=Row, column=3, columnspan=4, padx=5, pady=5, sticky=W)
		self.ShowColumnID.insert("end", 'StringId')

		Button(Tab, width = self.Button_Width_Full, text=  self.LanguagePack.Button['OpenOutput'], command= self.Open_BadWord_Result_Folder).grid(row=Row, column=9, columnspan=2,padx=5, pady=5, sticky=W)
		
		Row += 1
		Label(Tab, text="Search Value: ").grid(row=Row, column=1, padx=5, pady=5, sticky=W)
		self.ValueList = Text(Tab, width = 90, height=5) #
		self.ValueList.grid(row=Row, column=3, columnspan=4, padx=5, pady=5, sticky=W)
		'''
		Row += 1
		Label(Tab, text= self.LanguagePack.Label['MatchType']).grid(row=Row, column=1, padx=5, pady=5, sticky=W)
		self.Match_Type = IntVar()
		Radiobutton(Tab, width= 20, text=  self.LanguagePack.Option['ExactlyMatch'], value=1, variable=self.Match_Type).grid(row=Row, column=3, padx=0, pady=5, sticky=W)
		Radiobutton(Tab, width= 20, text=  self.LanguagePack.Option['Contains'], value=2, variable=self.Match_Type).grid(row=Row, column=5, padx=0, pady=5, sticky=W)
		self.Match_Type.set(1)

		Button(Tab, width = self.Button_Width_Full, text=  self.LanguagePack.Button['Execute'], command= self.Btn_BadWord_Execute).grid(row=Row, column=9, columnspan=2,padx=5, pady=5, sticky=W)


		Row += 1
		Label(Tab, text= self.LanguagePack.Label['Progress']).grid(row=Row, column=1, padx=5, pady=5, sticky=W)
		self.BadWord_Progressbar = Progressbar(Tab, orient=HORIZONTAL, length=850,  mode='determinate')
		self.BadWord_Progressbar["maximum"] = 1000
		self.BadWord_Progressbar.grid(row=Row, column=3, columnspan=9, padx=5, pady=5, sticky=W+E)
		
	
	def Generate_Language_Validator_UI(self, Tab):
		
		Row = 1
		Label(Tab, textvariable=self.Notice).grid(row=Row, column=1, columnspan = 10, padx=5, pady=5, sticky= W)
		Row += 1

		self.LanguageValidatorSource = StringVar()
		Label(Tab, text=  self.LanguagePack.Label['DataSource']).grid(row=Row, column=1, columnspan=2, padx=5, pady=5, sticky= W)


		self.TextLanguageSource = Entry(Tab, width = 120, state="readonly", textvariable=self.LanguageValidatorSource)
		self.TextLanguageSource.grid(row=Row, column=3, columnspan=4, padx=4, pady=5, sticky=W)

		Button(Tab, width = self.Button_Width_Full, text=  self.LanguagePack.Button['Browse'], command= self.Btn_Browse_Language_Validator_Data).grid(row=Row, column=9, columnspan=2, padx=5, pady=5, sticky=E)
		
		Row += 1
		Label(Tab, text="ColumnID: ").grid(row=Row, column=1, padx=5, pady=5, sticky=W)
		self.SourceColumnID = Text(Tab, width = 90, height=1) #
		self.SourceColumnID.grid(row=Row, column=3, columnspan=4, padx=5, pady=5, sticky=W)
		self.SourceColumnID.insert("end", 'Title ')

		Button(Tab, width = self.Button_Width_Full, text=  self.LanguagePack.Button['Execute'], command= self.Btn_Validate_Language).grid(row=Row, column=9, columnspan=2,padx=5, pady=5, sticky=W)


		Row += 1
		Label(Tab, text="Search Key: ").grid(row=Row, column=1, padx=5, pady=5, sticky=W)
		self.ShowColumnID = Text(Tab, width = 90, height=1) #
		self.ShowColumnID.grid(row=Row, column=3, columnspan=4, padx=5, pady=5, sticky=W)
		self.ShowColumnID.insert("end", 'StringId')

		Button(Tab, width = self.Button_Width_Full, text=  self.LanguagePack.Button['OpenOutput'], command= self.OpenValidateOutput).grid(row=Row, column=9, columnspan=2,padx=5, pady=5, sticky=W)

		Row += 1
		Label(Tab, text="Search Value: ").grid(row=Row, column=1, padx=5, pady=5, sticky=W)
		self.ValueList = Text(Tab, width = 90, height=5) #
		self.ValueList.grid(row=Row, column=3, columnspan=4, padx=5, pady=5, sticky=W)
	
	
	
	
	def Generate_Debugger_UI(self,Tab):
		Row = 1
		self.Debugger = Text(Tab, width=125, height=15, undo=True, wrap=WORD, )
		self.Debugger.grid(row=Row, column=1, columnspan=10, padx=5, pady=5, sticky=W+E+N+S)

###########################################################################################
# MENU FUNCTION
###########################################################################################

	def Menu_Function_About(self):
		messagebox.showinfo("About....", "Creator: Evan")

	def Show_Error_Message(self, ErrorText):
		messagebox.showinfo('Error...', ErrorText)	

	def SaveAppLanguage(self, language):

		self.Notice.set('Update app language...') 

		config = configparser.ConfigParser()
		config.read(self.AppConfig)
		if not config.has_section('DocumentToolkit'):
			config.add_section('DocumentToolkit')
			cfg = config['DocumentToolkit']	
		else:
			cfg = config['DocumentToolkit']

		cfg['applang']= language
		with open(self.AppConfig, 'w') as configfile:
			config.write(configfile)
		self.Notice.set('Config saved...')
		return

	def SetLanguageKorean(self):
		self.AppLanguage = 'kr'
		self.SaveAppLanguage(self.AppLanguage)
		#self.initUI()
	
	def SetLanguageEnglish(self):
		self.AppLanguage = 'en'
		self.SaveAppLanguage(self.AppLanguage)
		#self.initUI()

	def Function_Correct_Path(self, path):
		return str(path).replace('/', '\\')
	
	def Menu_Function_Open_Main_Guideline(self):
		webbrowser.open_new(r"https://confluence.nexon.com/display/NWMQA/Document+Toolkit")
	
	def Function_Correct_EXT(self, path, ext):
		if path != None and ext != None:
			Outputdir = os.path.dirname(path)
			baseName = os.path.basename(path)
			sourcename, Obs_ext = os.path.splitext(baseName)
			newPath = self.Function_Correct_Path(Outputdir + '/'+ sourcename + '.' + ext)
			return newPath

	def ErrorMsg(self, ErrorText):
		messagebox.showinfo('Error...', ErrorText)	

	

	def onExit(self):
		self.quit()


	def initGeneralSetting(self):
		
		config = configparser.ConfigParser()
		if os.path.isfile(self.AppConfig):
			config.read(self.AppConfig)
			if config.has_section('DocumentToolkit'):
				cfg = config['DocumentToolkit']
			else:
				config['DocumentToolkit'] = {}
				cfg = config['DocumentToolkit']

			if config.has_option('DocumentToolkit', 'applang'):	
				self.AppLanguage = config['DocumentToolkit']['applang']
				print('Setting saved: ', self.AppLanguage)
			else:
				self.AppLanguage = 'en'
				#print('Setting not saved: ', self.AppLanguage)

			#if config.has_option('Translator', 'Subscription'):
			#	self.Subscription = config['Translator']['Subscription']
			#else:
			#	self.Subscription = ''

		else:

			self.AppLanguage = 'en'

	def init_UI_Configuration(self):
		
		config = configparser.ConfigParser()
		if os.path.isfile(self.AppConfig):
			config.read(self.AppConfig)
			if config.has_section('Document_Utility'):
				cfg = config['Document_Utility']
			else:
				config['Document_Utility'] = {}
				cfg = config['Document_Utility']

			if config.has_section('Comparision'):
				cfg = config['Comparision']
			else:
				config.add_section('Comparision')
				cfg = config['Comparision']

		else:
			self.Language = 'en'
			

		return	

	def SaveSetting(self):

		print('Save setting')
		return

###########################################################################################
# OPTIMIZE DATA
###########################################################################################
	
	def Btn_Open_Optimizer_Folder(self):
		try:

			SourceDocument = self.Function_Correct_Path(self.OptimizedFolder)
			subprocess.Popen('explorer ' + "\"" + str(SourceDocument) + "\"")
		except AttributeError:
			self.ErrorMsg('Please select source folder.')
			return	
		BasePath = str(os.path.abspath(self.RawFile))
		subprocess.Popen('explorer ' + BasePath)
	
	def Btn_Load_Raw_Source(self):
		#filename = filedialog.askopenfilename(title =  self.LanguagePack.ToolTips['SelectSource'],filetypes = (("Workbook files","*.xlsx *.xlsm *xlsb"), ("Document files","*.docx")), multiple = True)	
		FolderName = filedialog.askdirectory(title =  self.LanguagePack.ToolTips['SelectSource'])	
		if FolderName != "":
			self.RawFile = FolderName
			self.RawSource.set(str(FolderName))

			Root = os.path.dirname(FolderName)

			Root_Name = os.path.basename(FolderName)
			Outputdir = Root + '//' + Root_Name + '_Optimized'
			
			if not os.path.isdir(Outputdir):
				try:
					os.mkdir(Outputdir)
				except OSError:
					Outputdir = FolderName + '//Optmized'

			self.OptimizedFolder = Outputdir
			
			self.Notice.set(self.LanguagePack.ToolTips['SourceSelected'])
		else:
			self.Notice.set(self.LanguagePack.ToolTips['SourceDocumentEmpty'])
		return



	def Btn_Optimize_XLSX(self):
		try:
			SourceDocument = self.RawFile
		except AttributeError:
			self.ErrorMsg('Please select source folder.')

		try:
			OutputDocument = self.OptimizedFolder
		except AttributeError:
			self.ErrorMsg('Please select source folder.')

		try:
			while True:
				percent = self.Process_Queue.get_nowait()
				#print("Remain percent: ", percent)
		except queue.Empty:
			pass
		self.Optimize_Process = Process(target=Function_Optimize_XLSX, args=(self.Status_Queue,self.Process_Queue, SourceDocument, OutputDocument, self.ExceptionList,))
		self.Optimize_Process.start()
		self.after(DELAY1, self.GetOptimizeStatus)	

	def GetOptimizeStatus(self):
		if (self.Optimize_Process.is_alive()):
			try:
				percent = self.Process_Queue.get(0)
				self.Optimize_Progressbar["value"] = percent
				self.Optimize_Progressbar.update()
				#self.Progress.set("Progress: " + str(percent/10) + '%')
			except queue.Empty:
				pass
			
			
			try:
				Status = self.Status_Queue.get(0)
				if Status != None:
					self.Notice.set(Status)
					self.Debugger.insert("end", "\n\r")
					self.Debugger.insert("end", Status)
					self.Debugger.yview(END)
			except queue.Empty:
				pass
			
			self.after(DELAY1, self.GetOptimizeStatus)
		else:
			try:
				Status = self.Status_Queue.get(0)
				if Status != None:	
					self.Notice.set(Status)
					self.Debugger.insert("end", "\n\r")
					self.Debugger.insert("end", Status)
					self.Debugger.yview(END)
			except queue.Empty:
				pass
			self.Optimize_Process.terminate()


###########################################################################################
# FAST COMPARISION
###########################################################################################
	
	def Open_Fast_Compare_Result_Folder(self):
		try:
			Source = self.FastCompare_NewDataTable[0]
			Outputdir = os.path.dirname(Source)
			BasePath = str(os.path.abspath(Outputdir)) + '\\Compare Result'
			subprocess.Popen('explorer ' + BasePath)
			print('SourceDocument:', BasePath)
		except AttributeError:
			self.ErrorMsg('Please select source folder.')
			return


	def BtnSelectColour(self):
		colorStr, self.BackgroundColor = colorchooser.askcolor(parent=self, title='Select Colour')
		
		
		if self.BackgroundColor == None:
			self.ErrorMsg('Set colour as defalt colour (Yellow)')
			self.BackgroundColor = 'ffff00'
		else:
			self.BackgroundColor = self.BackgroundColor.replace('#', '')
		#print(colorStr)
		#print(self.BackgroundColor)
		return


	def Btn_FastCompare_Browse_Old_Data_Folder(self):
		filename = filedialog.askopenfilename(title =  self.LanguagePack.ToolTips['SelectSource'],filetypes = (("Workbook files", "*.xlsx *.xlsm"), ), multiple = True)	
		if filename != "":
			self.FastCompare_OldDataTable = list(filename)
			self.Str_Old_File_Path.set(str(self.FastCompare_OldDataTable[0]))
			self.Notice.set(self.LanguagePack.ToolTips['SourceSelected'])	
		else:
			self.Notice.set(self.LanguagePack.ToolTips['SourceDocumentEmpty'])
		return

	def Btn_FastCompare_Browse_List_File(self):
		filename = filedialog.askopenfilename(title =  self.LanguagePack.ToolTips['SelectSource'],filetypes = (("Text files", "*.txt"), ), multiple = False)	
		if filename != "":
			self.List_Test_Files = filename
			self.Str_List_File_Test.set(str(filename))
			self.Notice.set(self.LanguagePack.ToolTips['SourceSelected'])	
		else:
			self.Notice.set(self.LanguagePack.ToolTips['SourceDocumentEmpty'])
		return	

	def Btn_FastCompare_Clear_Selected_File(self):
		self.List_Test_Files = ""
		self.Str_List_File_Test.set("")
		return	

	def Btn_FastCompare_Browse_New_Data_Folder(self):
		filename = filedialog.askopenfilename(title =  self.LanguagePack.ToolTips['SelectSource'],filetypes = (("Workbook files", "*.xlsx *.xlsm"), ), multiple = True)
		
		if filename != "":
			self.FastCompare_NewDataTable = list(filename)
			self.Str_New_File_Path.set(str(self.FastCompare_NewDataTable[0]))
			self.Notice.set(self.LanguagePack.ToolTips['SourceSelected'])	
		else:
			self.Notice.set(self.LanguagePack.ToolTips['SourceDocumentEmpty'])
		return

	
	def Btn_Fast_Compare_Data(self):

		try:
			OldDocument = self.FastCompare_OldDataTable
			NewDocument = self.FastCompare_NewDataTable
		except AttributeError:
			self.ErrorMsg('Please select source folder.')
			return

		try:
			Sheet_Name = self.Str_FastCompare_Data_Sheet_Name.get("1.0", END).replace('\n', '')
		except Exception as e:
			ErrorMsg = ('Error message: ' + str(e))
			print(ErrorMsg)

		Index_Col = ["Id"]
		try:
			Index_Col = self.Str_FastCompare_Data_Col_Name.get("1.0", END).replace('\n', '')
			Index_Col = Index_Col.replace(' ', '')
			Index_Col = Index_Col.split(',')
		except Exception as e:
			ErrorMsg = ('Error message: ' + str(e))
			print(ErrorMsg)
		

		try:
			self.Background_Color
		except:
			self.Background_Color = 'ffff00'	
		if self.Background_Color == False or self.Background_Color == None:
			self.Background_Color = 'ffff00'
		#print('self.BackgroundColor: ', self.BackgroundColor)

		try:
			self.Font_Color
		except:
			self.Font_Color = 'FF0000'	
		if self.Font_Color == False or self.Font_Color == None:
			self.Font_Color = 'FF0000'
		#print('Font_Color: ', self.Font_Color)
		#print('self.BackgroundColor: ', self.BackgroundColor)

		timestamp = Function_Get_TimeStamp()			
		path, filename = os.path.split(NewDocument[0])
		Output_Folder = path + '/Compare Result/' + 'Summary_Result_' + str(timestamp) + '.xlsx'

		Keep_Comment_Column = self.Skip_Comment_Column.get()

		list_files = []
		#self.List_Test_Files
		lines = []
		if (os.path.isfile(self.List_Test_Files)):
			with open(self.List_Test_Files) as f:
				lines = f.readlines()
		for line in lines:
			line = line.replace('\n', '')
			list_files.append(line)

		self.p4 = Process(target=Function_Fast_Compare_Data, args=(self.Status_Queue, self.Process_Queue, OldDocument, NewDocument, Output_Folder, Sheet_Name, Index_Col, Keep_Comment_Column, list_files, self.Background_Color, self.Font_Color,))
		self.p4.start()
		self.after(DELAY1, self.GetCompareStatus)	

	def GetCompareStatus(self):
		if (self.p4.is_alive()):
			try:
				Status = self.Status_Queue.get(0)
				if Status != None:
					self.Notice.set(Status)
					#print(Status)
					self.Debugger.insert("end", "\n\r")
					self.Debugger.insert("end", Status)
					self.Debugger.yview(END)
			except queue.Empty:
				pass	
			self.after(DELAY1, self.GetCompareStatus)
		else:
			try:
				Status = self.Status_Queue.get(0)
				if Status != None:	
					self.Notice.set(Status)
					#print(Status)
					self.Debugger.insert("end", "\n\r")
					self.Debugger.insert("end", Status)
					self.Debugger.yview(END)
			except queue.Empty:
				pass
			self.p4.terminate()


###########################################################################################
# PROFANITY DETECTOR 
###########################################################################################
	
	def Btn_BadWord_Select_Background_Colour(self):
		colorStr, self.Background_Color = colorchooser.askcolor(parent=self, title='Select Colour')
		
		if self.Background_Color == None:
			self.ErrorMsg('Set colour as defalt colour (Yellow)')
			self.Background_Color = 'ffff00'
		else:
			self.Background_Color = self.Background_Color.replace('#', '')
		#print(colorStr)
		#print(self.BackgroundColor)
		return
	
	def Btn_BadWord_Select_Font_Colour(self):
		colorStr, self.Font_Color = colorchooser.askcolor(parent=self, title='Select Colour')
		
		
		if self.Font_Color == None:
			self.ErrorMsg('Set colour as defalt colour (Yellow)')
			self.Font_Color = 'FF0000'
		else:
			self.Font_Color = self.Font_Color.replace('#', '')
		#print(colorStr)
		#print(self.BackgroundColor)
		return

	def Btn_BadWord_Browse_Data_Files(self):
			
		folder_name = filedialog.askdirectory(title =  self.LanguagePack.ToolTips['SelectSource'],)	
		if folder_name != "":
			self.BadWord_File_Path = folder_name
			self.Str_Text_File_Path.set(folder_name)
			self.Notice.set(self.LanguagePack.ToolTips['SourceSelected'])	
		else:
			self.Notice.set(self.LanguagePack.ToolTips['SourceDocumentEmpty'])

	def Btn_BadWord_Browse_Data_Files_Old(self):
			
		filename = filedialog.askopenfilename(title =  self.LanguagePack.ToolTips['SelectSource'], filetypes = (("Workbook files", "*.xlsx *.xlsm"), ), multiple = True)	
		if filename != "":
			self.BadWord_File_Path = list(filename)
			self.Str_Text_File_Path.set(str(self.BadWord_File_Path[0]))
			
			self.Notice.set(self.LanguagePack.ToolTips['SourceSelected'])	
		else:
			self.Notice.set(self.LanguagePack.ToolTips['SourceDocumentEmpty'])
		return

	def Btn_BadWord_Browse_DB_File(self):
			
		filename = filedialog.askopenfilename(title =  self.LanguagePack.ToolTips['SelectSource'], filetypes = (("Workbook files", "*.xlsx *.xlsm"), ), multiple = False)	
		if filename != "":
			self.BadWord_DB_Path = filename
			self.Str_BadWord_DB_Path.set(filename)
			self.Notice.set(self.LanguagePack.ToolTips['SourceSelected'])	
		else:
			self.Notice.set(self.LanguagePack.ToolTips['SourceDocumentEmpty'])
		return
		
	def Open_BadWord_Result_Folder(self):
		try:
			subprocess.Popen('explorer ' + self.BadWord_File_Path)
		except AttributeError:
			self.ErrorMsg('Please select source folder.')
			return


	def Btn_BadWord_Execute(self):
		Text_Folder = self.BadWord_File_Path
		_temp_text_files = os.listdir(Text_Folder)
		Text_Files = []
		for file in _temp_text_files:
			file_path = Text_Folder + '/' + file
			if os.path.isfile(file_path):
				baseName = os.path.basename(file_path)
				sourcename, ext = os.path.splitext(baseName)
				if 'xls' in ext:
					Text_Files.append(file_path)

		match_type_index = self.Match_Type.get()
		if match_type_index == 1:
			exact_match = True
		else:
			exact_match = False

		Db_File = self.BadWord_DB_Path

		Sheet_Name = "Data"
		
		try:
			Sheet_Name = self.BadWord_Data_Sheet_Name.get("1.0", END).replace('\n', '')
		except Exception as e:
			ErrorMsg = ('Error message: ' + str(e))
			print(ErrorMsg)

		Index_Col = "String"
		try:
			Index_Col = self.BadWord_ColumnID.get("1.0", END).replace('\n', '')
		except Exception as e:
			ErrorMsg = ('Error message: ' + str(e))
			print(ErrorMsg)
		

		try:
			self.Background_Color
		except:
			self.Background_Color = 'ffff00'	
		if self.Background_Color == False or self.Background_Color == None:
			self.Background_Color = 'ffff00'
		#print('Background_Color: ', self.Background_Color)
		
		try:
			self.Font_Color
		except:
			self.Font_Color = 'FF0000'	
		if self.Font_Color == False or self.Font_Color == None:
			self.Font_Color = 'FF0000'
		#print('Font_Color: ', self.Font_Color)

		timestamp = Function_Get_TimeStamp()			
		Output_Result_Folder = Text_Folder + '/' + 'Bad_Word_Result_' + str(timestamp)
		if not os.path.isdir(Output_Result_Folder):
			os.mkdir(Output_Result_Folder)
			
		self.BadWord_Check_Process = Process(target=Function_BadWord_Execute, args=(self.Status_Queue, self.Process_Queue, Text_Files, Db_File, Output_Result_Folder, Sheet_Name, Index_Col, exact_match, self.Background_Color, self.Font_Color,))
		self.BadWord_Check_Process.start()
		self.after(DELAY1, self.Wait_For_BadWord_Process)	

	def Wait_For_BadWord_Process(self):
		if (self.BadWord_Check_Process.is_alive()):
			
			try:
				percent = self.Process_Queue.get(0)
				self.BadWord_Progressbar["value"] = percent
				self.BadWord_Progressbar.update()
				#self.Progress.set("Progress: " + str(percent/10) + '%')
			except queue.Empty:
				pass	
			
			try:
				Status = self.Status_Queue.get(0)
				if Status != None:
					self.Notice.set(Status)
					self.Debugger.insert("end", "\n\r")
					self.Debugger.insert("end", Status)
					self.Debugger.yview(END)
			except queue.Empty:
				pass	
			self.after(DELAY1, self.Wait_For_BadWord_Process)
		else:
			try:
				percent = self.Process_Queue.get(0)
				self.BadWord_Progressbar["value"] = percent
				self.BadWord_Progressbar.update()
				#self.Progress.set("Progress: " + str(percent/10) + '%')
			except queue.Empty:
				pass
			try:
				Status = self.Status_Queue.get(0)
				if Status != None:	
					self.Notice.set('Bad word check is completed')
					#print(Status)
					self.Debugger.insert("end", "\n\r")
					self.Debugger.insert("end", Status)
					self.Debugger.yview(END)
			except queue.Empty:
				pass
			self.BadWord_Check_Process.terminate()

	def ImportException(self, ExceptionPath):
		print('Loading My Exception list:', ExceptionPath)
		if ExceptionPath != None:
			if (os.path.isfile(ExceptionPath)):
				
				xlsx = load_workbook(ExceptionPath)	
				Exception = []
				for sheet in xlsx:
					database = None
					TextColl = ""
					ws = xlsx[sheet.title]
					for row in ws.iter_rows():
						
						for cell in row:
							text = str(cell.value).lower()
							if text == "exception":
								TextColl = cell.column_letter
								Row = cell.row
								database = ws
								break	
						if database != None:
							break	

					if database != None:
						for i in range(Row, database.max_row): 
							ExceptionAdderss = TextColl + str(i+1)
							ExceptionCell = database[ExceptionAdderss]
							ExceptionValue = ExceptionCell.value
							if ExceptionValue == None:
								continue
							else:
								Exception.append(ExceptionValue)
						return Exception
					else:
						return []
			else:
				print('Exception list is not existed')
				return []
		else:
			return []

###########################################################################################
# DEEP COMPARISION
###########################################################################################
	def Btn_Select_Background_Colour(self):
		colorStr, self.Background_Color = colorchooser.askcolor(parent=self, title='Select Colour')
		
		
		if self.Background_Color == None:
			self.ErrorMsg('Set colour as defalt colour (Yellow)')
			self.Background_Color = 'ffff00'
		else:
			self.Background_Color = self.Background_Color.replace('#', '')
		#print(colorStr)
		#print(self.BackgroundColor)
		return
	
	def Btn_Select_Font_Colour(self):
		colorStr, self.Font_Color = colorchooser.askcolor(parent=self, title='Select Colour')
		
		
		if self.Font_Color == None:
			self.ErrorMsg('Set colour as defalt colour (Yellow)')
			self.Font_Color = 'FF0000'
		else:
			self.Font_Color = self.Font_Color.replace('#', '')
		#print(colorStr)
		#print(self.BackgroundColor)
		return



	def Btn_Browse_Old_Data_File(self):
			
		filename = filedialog.askopenfilename(title =  self.LanguagePack.ToolTips['SelectSource'],filetypes = (("Workbook files", "*.xlsx *.xlsm"), ), multiple = True)	
		if filename != "":
			self.Old_File_Path = list(filename)
			self.Str_Deep_Old_File_Path.set(str(self.Old_File_Path[0]))
			
			self.Notice.set(self.LanguagePack.ToolTips['SourceSelected'])	
		else:
			self.Notice.set(self.LanguagePack.ToolTips['SourceDocumentEmpty'])
		return

	def Open_Deep_Compare_Result_Folder(self):
		try:
			Source = self.New_File_Path[0]
			Outputdir = os.path.dirname(Source)
			BasePath = str(os.path.abspath(Outputdir)) + '\\Compare Result'
			subprocess.Popen('explorer ' + BasePath)
			print('SourceDocument:', BasePath)
		except AttributeError:
			self.ErrorMsg('Please select source folder.')
			return

	def Btn_Browse_New_Data_File(self):
		
		filename = filedialog.askopenfilename(title =  self.LanguagePack.ToolTips['SelectSource'],filetypes = (("Workbook files", "*.xlsx *.xlsm"), ), multiple = True)
		
		if filename != "":
			self.New_File_Path = list(filename)
			self.Str_Deep_New_File_Path.set(str(self.New_File_Path[0]))
			self.Notice.set(self.LanguagePack.ToolTips['SourceSelected'])	
		else:
			self.Notice.set(self.LanguagePack.ToolTips['SourceDocumentEmpty'])
		return

	def Btn_Deep_Compare_Data(self):
		Old_File = self.Old_File_Path
		New_File = self.New_File_Path

		Sheet_Name = "Data"
		
		try:
			Sheet_Name = self.Str_Data_Sheet_Name.get("1.0", END).replace('\n', '')
		except Exception as e:
			ErrorMsg = ('Error message: ' + str(e))
			print(ErrorMsg)

		Index_Col = "Id"
		try:
			Index_Col = self.Str_Data_Col_Name.get("1.0", END).replace('\n', '')
		except Exception as e:
			ErrorMsg = ('Error message: ' + str(e))
			print(ErrorMsg)
		

		try:
			self.Background_Color
		except:
			self.Background_Color = 'ffff00'	
		if self.Background_Color == False or self.Background_Color == None:
			self.Background_Color = 'ffff00'
		#print('Background_Color: ', self.Background_Color)
		
		try:
			self.Font_Color
		except:
			self.Font_Color = 'FF0000'	
		if self.Font_Color == False or self.Font_Color == None:
			self.Font_Color = 'FF0000'
		#print('Font_Color: ', self.Font_Color)

		timestamp = Function_Get_TimeStamp()			
		path, filename = os.path.split(New_File[0])
		Output_Result = path + '/' + 'Compare_Result_' + str(timestamp) + '.xlsx'

		self.Data_Compare_Process = Process(target=Function_Deep_Compare_Data, args=(self.Status_Queue, self.Process_Queue, Old_File, New_File, self.ExceptionList, Output_Result, Sheet_Name, Index_Col, self.Background_Color, self.Font_Color,))
		self.Data_Compare_Process.start()
		self.after(DELAY1, self.Wait_For_Data_Compare_Process)	

	def Wait_For_Data_Compare_Process(self):
		if (self.Data_Compare_Process.is_alive()):
			'''
			try:
				percent = self.ProcessQueue.get(0)
				self.CompareProgressbar["value"] = percent
				self.progressbar.update()
				#self.Progress.set("Progress: " + str(percent/10) + '%')
			except queue.Empty:
				pass	
			'''
			try:
				Status = self.Status_Queue.get(0)
				if Status != None:
					self.Notice.set(Status)
					self.Debugger.insert("end", "\n\r")
					self.Debugger.insert("end", Status)
					self.Debugger.yview(END)
			except queue.Empty:
				pass	
			self.after(DELAY1, self.Wait_For_Data_Compare_Process)
		else:
			try:
				Status = self.Status_Queue.get(0)
				if Status != None:	
					self.Notice.set('Compare complete')
					#print(Status)
					self.Debugger.insert("end", "\n\r")
					self.Debugger.insert("end", Status)
					self.Debugger.yview(END)
			except queue.Empty:
				pass
			self.Data_Compare_Process.terminate()

	def ImportException(self, ExceptionPath):
		print('Loading My Exception list:', ExceptionPath)
		if ExceptionPath != None:
			if (os.path.isfile(ExceptionPath)):
				
				xlsx = load_workbook(ExceptionPath)	
				Exception = []
				for sheet in xlsx:
					database = None
					TextColl = ""
					ws = xlsx[sheet.title]
					for row in ws.iter_rows():
						
						for cell in row:
							text = str(cell.value).lower()
							if text == "exception":
								TextColl = cell.column_letter
								Row = cell.row
								database = ws
								break	
						if database != None:
							break	

					if database != None:
						for i in range(Row, database.max_row): 
							ExceptionAdderss = TextColl + str(i+1)
							ExceptionCell = database[ExceptionAdderss]
							ExceptionValue = ExceptionCell.value
							if ExceptionValue == None:
								continue
							else:
								Exception.append(ExceptionValue)
						return Exception
					else:
						return []
			else:
				print('Exception list is not existed')
				return []
		else:
			return []

###########################################################################################
# Auto test
###########################################################################################
	# Not use
	def Btn_AutoTest_Select_Background_Colour(self):
		colorStr, self.Background_Color = colorchooser.askcolor(parent=self, title='Select Colour')
		
		if self.Background_Color == None:
			self.ErrorMsg('Set colour as defalt colour (Yellow)')
			self.Background_Color = 'ffff00'
		else:
			self.Background_Color = self.Background_Color.replace('#', '')
		#print(colorStr)
		#print(self.BackgroundColor)
		return
	
	# Not use
	def Btn_AutoTest_Select_Font_Colour(self):
		colorStr, self.Font_Color = colorchooser.askcolor(parent=self, title='Select Colour')
			
		if self.Font_Color == None:
			self.ErrorMsg('Set colour as defalt colour (Yellow)')
			self.Font_Color = 'FF0000'
		else:
			self.Font_Color = self.Font_Color.replace('#', '')
		#print(colorStr)
		#print(self.BackgroundColor)
		return

	def Btn_AutoTest_Browse_Old_Folder(self):
			
		folder_name = filedialog.askdirectory(title =  self.LanguagePack.ToolTips['SelectSource'],)	
		if folder_name != "":
			self.AutoTest_Old_Folder_Path = folder_name
			self.Str_AutoTest_Old_Folder_Path.set(folder_name)
			self.Notice.set(self.LanguagePack.ToolTips['SourceSelected'])	
		else:
			self.Notice.set(self.LanguagePack.ToolTips['SourceDocumentEmpty'])


	def Btn_AutoTest_Browse_New_Folder(self):
		
		folder_name = filedialog.askdirectory(title =  self.LanguagePack.ToolTips['SelectSource'],)
		
		if folder_name != "":
			self.AutoTest_New_Folder_Path = folder_name
			self.Str_AutoTest_New_Folder_Path.set(str(folder_name))
			self.Notice.set(self.LanguagePack.ToolTips['SourceSelected'])	
		else:
			self.Notice.set(self.LanguagePack.ToolTips['SourceDocumentEmpty'])

	
	def Btn_AutoTest_Browse_Test_Config(self):
	
		filename = filedialog.askopenfilename(title =  self.LanguagePack.ToolTips['SelectSource'],filetypes = (("Workbook files", "*.xlsx *.xlsm"), ), multiple = False)
		
		if filename != "":
			self.test_configuration_file = filename
			self.Str_Test_Configuration_File_Path.set(filename)
			self.Notice.set(self.LanguagePack.ToolTips['SourceSelected'])	
		else:
			self.Notice.set(self.LanguagePack.ToolTips['SourceDocumentEmpty'])
		return

	def Btn_AutoTest_Clear_Selected_File(self):
		self.List_Test_Files = ""
		self.Str_List_File_Test.set("")
		return	

	def Btn_AutoTest_Execute(self):
		Old_Folder = self.AutoTest_Old_Folder_Path
		New_Folder = self.AutoTest_New_Folder_Path

		try:
			self.Background_Color
		except:
			self.Background_Color = 'ffff00'	
		if self.Background_Color == False or self.Background_Color == None:
			self.Background_Color = 'ffff00'
		#print('Background_Color: ', self.Background_Color)
		
		try:
			self.Font_Color
		except:
			self.Font_Color = 'FF0000'	
		if self.Font_Color == False or self.Font_Color == None:
			self.Font_Color = 'FF0000'
		#print('Font_Color: ', self.Font_Color)

		timestamp = Function_Get_TimeStamp()			
		#Output_Result = New_Folder + '\\' + 'Autotest_Result_' + str(timestamp) + '.xlsx'

		self.AutoTest_Process = Process(target=Function_AutoTest, args=(self.Status_Queue, self.Process_Queue, self.test_configuration_file, Old_Folder, New_Folder, self.Background_Color, self.Font_Color,))
		self.AutoTest_Process.start()
		self.after(DELAY1, self.Wait_For_AutoTest_Process)	

	def Wait_For_AutoTest_Process(self):
		if (self.AutoTest_Process.is_alive()):
	
			try:
				Status = self.Status_Queue.get(0)
				if Status != None:
					self.Notice.set(Status)
					self.Debugger.insert("end", "\n\r")
					self.Debugger.insert("end", Status)
					self.Debugger.yview(END)
			except queue.Empty:
				pass	

			try:
				percent = self.Process_Queue.get(0)
				self.progressbar["value"] = percent
				self.progressbar.update()
				self.Progress.set("Progress: " + str(percent/10) + '%')
			except queue.Empty:
				pass



			self.after(DELAY1, self.Wait_For_AutoTest_Process)
		else:
			try:
				Status = self.Status_Queue.get(0)
				if Status != None:	
					self.Notice.set('Compare complete')
					#print(Status)
					self.Debugger.insert("end", "\n\r")
					self.Debugger.insert("end", Status)
					self.Debugger.yview(END)
			except queue.Empty:
				pass
			self.AutoTest_Process.terminate()

	def Btn_Open_AutoTest_Result_Folder(self):
		try:
			Source = self.Function_Correct_Path(self.AutoTest_New_Folder_Path)
			
			print('explorer ' + "\"" + Source + "\"")
			subprocess.Popen('explorer ' + "\"" + str(Source) + "\"")
		except AttributeError:
			self.ErrorMsg('Please select source folder.')
			return
###########################################################################################

def Function_Optimize_XLSX( StatusQueue, ProgressQueue, SourceDocument, OutputDocument, ExceptionList,):

	FileList = os.listdir(SourceDocument)

	TotalFile = len(FileList)
	Complete = 0

	TaskList = []	
	processes = []
	number_of_processes = multiprocessing.cpu_count()

	for FileName in FileList:
		str_filename = str(FileName)
		if FileName != None and FileName not in ExceptionList:
			File = SourceDocument + '//' + FileName
			TaskList.append(File)

		else:
			StatusQueue.put('Skip, ' + str_filename + ' is in the exception list.')
			Complete+=1
			percent = ShowProgress(Complete, TotalFile)
			ProgressQueue.put(percent)

	while len(TaskList) > 0:
		if len(TaskList) > number_of_processes:
			NewTask = number_of_processes
		else:
			NewTask = len(TaskList)

		for w in range(NewTask):
			input_file = TaskList[0]

			baseName = os.path.basename(input_file)
			sourcename, ext = os.path.splitext(baseName)
			output_file = OutputDocument + '//' + sourcename + ext
			StatusQueue.put('Process file: ' + sourcename)	
			p = Process(target=Optimize_Single_File, args=(input_file, output_file,))

			del TaskList[0]
			processes.append(p)
			p.start()

		for p in processes :
			p.join()
			Complete+=1
		
		percent = ShowProgress(Complete, TotalFile)
		ProgressQueue.put(percent)
		
	StatusQueue.put('Optimized done.')	

def Optimize_Single_File( InputDocument, OutputDocument):

	try:
		xlsx = load_workbook(InputDocument, data_only=True)
	except:
		print('Error')
	#xlsx.active = 1
	for sheet in xlsx.get_sheet_names():
		if sheet != 'Data':
			std=xlsx.get_sheet_by_name(sheet)
			xlsx.remove_sheet(std)
	try:
		xlsx.save(OutputDocument)
	except Exception as e:
		print('Error')
	xlsx.close()

###########################################################################################

def Function_Generate_DB( StatusQueue, ProgressQueue, SourceDocument, OutputDocument,):
	from openpyxl import load_workbook, worksheet, Workbook
	from openpyxl.styles import Font

	FileList = os.listdir(SourceDocument)

	TotalFile = len(FileList)
	Complete = 0

	TaskList = []	
	processes = []
	number_of_processes = multiprocessing.cpu_count()

	for FileName in FileList:
		str_filename = str(FileName)
		if FileName != None and "String" in FileName:
			File = SourceDocument + '//' + FileName
			TaskList.append(File)

	while len(TaskList) > 0:
		
		#Import_Data
		#Append DB
		
		del TaskList[0]

	StatusQueue.put('Optimized done.')	

def Get_DB_Info( InputDocument, OutputDocument):

	try:
		xlsx = load_workbook(InputDocument, data_only=True)
	except:
		return False
	
	try:
		xlsx.save(OutputDocument)
	except Exception as e:
		return False

###########################################################################################

'''
class DataProcessing(Test)
	def __init__(self, SourceDocument, Main_Sheet="Data"):
		self.FileList = os.listdir(SourceDocument)
		self.Main = Main_Sheet
		self.DF = []
		self.Init()

	def Init(self):
		for FileName in FileList:
			if FileName != None:
				File = SourceDocument + '//' + FileName
				Current_DF = {}
				Current_DF.File = FileName
				
				try:
					Current_DF.DF = pd.read_excel(File, sheet_name="Data")	
				except:
					continue
				self.DF.append(Current_DF)

	self.Data_Lookup(Lookup_Value):

		return	

def Function_Lookup_Data(SourceDocument, StatusQueue, ProgressQueue):
	from openpyxl import load_workbook, worksheet, Workbook
	from openpyxl.styles import Font
	#print(SourceDocument)
	FileList = os.listdir(SourceDocument)
	TotalFile = len(FileList)
	Complete = 0
	#print(FileList)
	#Outputdir = os.path.dirname(File) + '/Optimized/'
	FolderName = os.path.basename(SourceDocument)
	Outputdir = SourceDocument + '//' + FolderName + '_Optimized'
	if not os.path.isdir(Outputdir):
		try:
			os.mkdir(Outputdir)
		except OSError:
			#print ("Creation of the directory %s failed" % Outputdir)
			Outputdir = SourceDocument + '//Optmized'

	AllData = {}

	for FileName in FileList:
		if FileName != None:
			File = SourceDocument + '//' + FileName
			try:
				Data = pd.read_excel(File, sheet_name="Data")	
			except:
				continue
			
			baseName = os.path.basename(File)
			sourcename, ext = os.path.splitext(baseName)
			
			StatusQueue.put('Optimizing file: ' + sourcename)
			
			output_file = Outputdir+ '//' + sourcename + ext
			#print('output_file:',output_file)
			
			try:
				xlsx.save(output_file)
				#StatusQueue.put('Optimized done.')
			except Exception as e:
				StatusQueue.put('Failed to save the result: ' + str(e))
			
			Complete+=1
			
			percent = ShowProgress(Complete, TotalFile)
			ProgressQueue.put(percent)				

	StatusQueue.put('Optimized done.')	
'''
###########################################################################################
def ShowProgress(Counter, TotalProcess):
	#os.system('CLS') 
	percent = int(1000 * Counter / TotalProcess)
	#print("Current progress: " +  str(Counter) + '/ ' + str(TotalProcess))
	return percent

###########################################################################################
#
def Function_Fast_Compare_Single_Data(Status_Queue, Old_File, New_File, Sheet_Name, Index_Col_Name, result_filename = None, Keep_Comment_Column = True,  Background_Colour = 'ffff00', Font_Colour = 'FF0000',):
	BaseName = os.path.basename(Old_File)
	Result = {}
	Status_Queue.put("Import new data: " + New_File)
	New_Data, New_List_Col, New_Index_ID = Function_Fast_Create_Data_TC(New_File, Sheet_Name, Index_Col_Name, Keep_Comment_Column)
	
	Status_Queue.put("Import old data: " + Old_File)
	Old_Data, Old_List_Col, Old_Index_ID = Function_Fast_Create_Data_TC(Old_File, Sheet_Name, Index_Col_Name, Keep_Comment_Column)
	
	merged_list_col = Old_List_Col
	for item in New_List_Col:
		if item not in Old_List_Col:
			merged_list_col.append(item)

	Status_Queue.put("Compare data for " + BaseName)

	Old_Data['version'] = "old"
	New_Data['version'] = "new"	

	old_accts_all = set(Old_Data[Old_Index_ID])
	new_accts_all = set(New_Data[New_Index_ID])

	dropped_accts = old_accts_all - new_accts_all

	added_accts = new_accts_all - old_accts_all

	all_data = pd.concat([Old_Data,New_Data], ignore_index=True)

	changes = all_data.drop_duplicates(subset=None, keep= 'last')

	dupe_accts = changes[changes[Old_Index_ID].duplicated() == True][Old_Index_ID].tolist()
	dupes = changes[changes[Old_Index_ID].isin(dupe_accts)]
	
	change_new = dupes[(dupes["version"] == "new")]
	change_old = dupes[(dupes["version"] == "old")]
	change_new = change_new.drop(['version'], axis=1)
	change_old = change_old.drop(['version'], axis=1)

	change_new.set_index(Old_Index_ID, inplace=True)
	change_new = change_new.fillna("#NA")

	change_old.set_index(Old_Index_ID, inplace=True)
	change_old = change_old.fillna("#NA")

	try:	
		df_all_changes = pd.concat([change_old, change_new],
									axis='columns',
									keys=['old', 'new'],
									join='outer')
		df_all_changes = df_all_changes.swaplevel(axis='columns')[change_new.columns[0:]]
		print('All', len(df_all_changes))
		# Compare each column and modify the current value of the cell.
		df_changed = df_all_changes.groupby(level=0, axis=1).apply(lambda frame: frame.apply(report_diff, axis=1))
		
		df_changed = df_changed.reset_index()	
		# Check if the row has the text -->
		pd.set_option("display.max_colwidth", None)

		#for column in df_changed:
		#	df_changed['has_change'] = df_changed[column].apply(cell_has_change)

		print('Changed', len(df_changed))
		df_changed['has_change'] = df_changed.apply(cell_has_change, axis=1)
		diff = df_changed[(df_changed['has_change'] == 'Y')]	
		print('Diff', len(diff))
		diff = diff.reindex(columns=merged_list_col)
		

	except:
		diff = []

	dropped = changes[changes[Old_Index_ID].isin(dropped_accts)]
	
	dropped = dropped.drop(['version'], axis=1)
	dropped.set_index(Old_Index_ID,inplace=True)
	dropped = dropped.reset_index()
	dropped = dropped.dropna(how='all')
	
	added = changes[changes[Old_Index_ID].isin(added_accts)]

	added = added.drop(['version'], axis=1)
	added.set_index(Old_Index_ID,inplace=True)
	added = added.reset_index()
	added = added.dropna(how='all')
	
	if result_filename == None:
		result_filename = Function_Add_Surflix(New_File, 'Compare Result', "Fast_Compare")

	wb = Workbook()
	ws =  wb.active
	ws.title = 'Diff'
	changed = False
	if len(diff) > 0:
		changed = True
		ws = wb['Diff']
		ws = print_worksheet(ws, diff, 'diff', Font_Colour, Background_Colour, mark_changes = True)
	if len(added) > 0:
		changed = True
		wb.create_sheet('Added')
		ws = wb['Added']
		ws = print_worksheet(ws, added, 'added', Font_Colour, Background_Colour)
	if len(dropped) > 0:
		changed = True
		wb.create_sheet('Dropped')
		ws = wb['Dropped']
		ws = print_worksheet(ws, dropped, 'dropped', Font_Colour, Background_Colour)
	
	if changed == True:
		if len(diff) == 0:
			wb.remove(wb['Diff'])
		try:	
			wb.save(result_filename)
		except:
			Status_Queue.put("Permission denied, fail to save result file!")
		Status_Queue.put("Changed, please check result file.")
		Result['Changed'] = True
		Result['Type'] = 'Changed'
		Result['Details'] = {}
		Result['Details']['Dropped'] = len(dropped)
		Result['Details']['Added'] = len(added)
		Result['Details']['Diff'] = len(diff)

	else:
		Status_Queue.put("No change")
		Result['Changed'] = False
		Result['Type'] = 'No change'	

	return Result

def Function_Fast_Compare_Data(	
		Status_Queue, Process_Queue, Old_Files, New_Files, out_path, Sheet_Name, Index_Col_Name, Keep_Comment_Column = True, list_files = [], Background_Colour = 'ffff00', Font_Colour = 'FF0000', **kwargs):
	my_color = Color(rgb=Background_Colour)
	my_fill = PatternFill(patternType='solid', fgColor=my_color)
	my_font = Font(color=Font_Colour)
	if isinstance(Old_Files, str):
		Old_Files = [Old_Files]
	if isinstance(New_Files, str):
		New_Files = [New_Files]
	if isinstance(Index_Col_Name, str):
		Index_Col_Name = [Index_Col_Name]	
	Start = time.time()

	Sum_Result = {}

	main_index = ''
	for index in Index_Col_Name:
		if index == Index_Col_Name[0]:
			main_index += index
		else:
			main_index += '.' + index
	
	Old_Outputdir = os.path.dirname(Old_Files[0])

	New_Outputdir = os.path.dirname(New_Files[0])

	for File in list_files:
		old_file = Old_Outputdir + '/' + File
		if check_exist_file(old_file):
			Old_Files.append(old_file)
		new_file = New_Outputdir + '/' + File
		if check_exist_file(new_file):
			New_Files.append(new_file)

	for Old_File in Old_Files:
		Old_Outputdir = os.path.dirname(Old_File)
		Old_baseName = os.path.basename(Old_File)
		Old_sourcename, Old_ext = os.path.splitext(Old_baseName)

		Sum_Result[Old_baseName] = {}
		Result = Sum_Result[Old_baseName]
		Result['Changed'] = False
		Found = False
		for New_File in New_Files: 	
			New_baseName = os.path.basename(New_File)
			
			if Old_baseName == New_baseName:
				Found = True	
				Compare_Result = Function_Fast_Compare_Single_Data(Status_Queue, Old_File, New_File, Sheet_Name, Index_Col_Name, None, True, Background_Colour, Font_Colour,)
				Sum_Result[Old_baseName] = Compare_Result
		if not Found:
			Status_Queue.put("File dropped")
			Result['Changed'] = True
			Result['Type'] = 'Dropped'
			

	for New_File in New_Files:
		New_baseName = os.path.basename(New_File)

		if New_baseName in Sum_Result:
			continue
		else:
			Status_Queue.put("File added")
			Sum_Result[New_baseName] = {}
			Result = Sum_Result[New_baseName]
			Result['Changed'] = False
			Result['Type'] = 'Added'
	
	summary = Workbook()
	ws =  summary.active
	ws.title = 'Summary'
	Header = ['File', 'Result', 'Changed', 'Added', 'Dropped']
	Col = 2
	Row = 2
	for Par in Header:
		ws.cell(row=Row, column=Col).value = Par
		Col +=1
	Row +=1
	column_letters = ['B', 'C', 'D', 'E', 'F']

	for column_letter in column_letters:
		ws.column_dimensions[column_letter].bestFit = True
	print('Sum_Result', Sum_Result)
	for file_name in Sum_Result:
		ws.cell(row=Row, column=2).value = file_name
		Changed = Sum_Result[file_name]['Changed']
		ws.cell(row=Row, column=3).value = Sum_Result[file_name]['Type']
		if Changed:
			if 'Details' in Sum_Result[file_name]:
				ws.cell(row=Row, column=4).value = Sum_Result[file_name]['Details']['Diff']
				ws.cell(row=Row, column=5).value = Sum_Result[file_name]['Details']['Added']
				ws.cell(row=Row, column=6).value = Sum_Result[file_name]['Details']['Dropped']
		Row +=1
	
	Tab = Table(displayName="Summary", ref="B2:" + "F" + str(Row-1))
	style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=True)
	Tab.tableStyleInfo = style
	ws.add_table(Tab)
	now = datetime.now()
	timestamp = str(int(datetime.timestamp(now)))	

	summary.save(out_path)	
	summary.close()

	Status_Queue.put("Result found at: " + str(out_path))
	End = time.time()
	Total = End - Start
	Status_Queue.put("'Time spent: " + str(Total))
	print('Time spent:', Total)	
	print ('\nDone.\n')


def check_exist_file(path):
	if (os.path.isfile(path)):
		return True
	return False	


def print_worksheet(worksheet, dataframe,TabName, Font_Color, Background_Color, mark_changes = False):
	for r in dataframe_to_rows(dataframe, index=False, header=True):
		try:
			worksheet.append(r)
		except Exception as e:
			print('Error when append row: ', e)

	
	if mark_changes:
		my_font = Font(color=Font_Color)	
		my_fill = PatternFill(patternType='solid', fgColor=Background_Color)
		for row in worksheet.iter_rows():
			for cell in row:
				cell_string = str(cell.value)
				if "-->" in cell_string:
					cell.fill = my_fill
					cell.font = my_font
	
	'''
	for cell in worksheet['A'] + worksheet[1]:
		cell.style = 'Pandas'
	'''

	#row_count_str=str(len(dataframe.index)+1)
	diff_row = worksheet.max_row+1 
	diff_col = Get_Column(worksheet.max_column)
	


	Tab = Table(displayName=TabName, ref="A1:" + diff_col + str(diff_row-1))
	style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=True)
	Tab.tableStyleInfo = style
	worksheet.add_table(Tab)

	for col in worksheet.columns:
		max_length = 0
		column = col[0].column_letter # Get the column name
	# Since Openpyxl 2.6, the column name is  ".column_letter" as .column became the column number (1-based) 
		for cell in col:
			try: # Necessary to avoid error on empty cells
				if len(str(cell.value)) > max_length:
					max_length = len(cell.value)
			except:
				pass
		adjusted_width = (max_length + 2) * 1.2
		worksheet.column_dimensions[column].width = adjusted_width
	

	return worksheet

def Get_Column(number):

	letter = ""
	while (number > 0):
		modulo = (number - 1) % 26
		letter = chr(65 + modulo).upper() + letter
		number = (int)((number - modulo) / 26)

	return letter

def has_change(row):
	#print(row.to_string())
	for cell in row:
		if "-->" in str(cell):
			#print(cell)
			return 'Y' 
		else:
			print(cell)	
	return "N"

def cell_has_change(cell):
	if "-->" in str(cell.to_string()):
		return "Y"
	else:
		return "N"

def Drop_Empty(df):
	empty_cols = [col for col in df.columns if df[col].isnull().all()]
	# Drop these columns from the dataframe
	df.drop(empty_cols,
			axis=1,
			inplace=True)
	return df



def report_diff(x):
	#print(x)
	if len(x) == 2:
		if x[0] == x[1]:
			#return x[0]
			return x[0]
		elif x[0] == "#NA":
			return "[ADD] --> " + str(x[1])
		elif x[1]== "#NA":
			return "[DROP] --> " + str(x[0])
		else:
			#print('diff found','\n', x[0],'\n', x[1])
			return str(x[0]) + " --> " + str(x[1])
			#return '{} --> {}'.format(*x)
	else:
		#return x[0]
		return x[0]

def Function_Fast_Create_Data_TC(Data_Workbook, Sheet_Name, Index_Col, Keep_Comment_Column = True):
	#print('Fast create data')
	#print('Source:', Data_Workbook)
	List_Col = []
	default_index = Index_Col[0]
	#print('default_index', default_index)
	#print('Function_Fast_Create_Data', 'Index_Col', Index_Col)
	#print('Load workbook')
	xlsx = load_workbook(Data_Workbook, data_only = True)	
	#print('Load col list', Index_Col)
	ws = xlsx[Sheet_Name]
	#print('Get Column Row')
	for row in ws.iter_rows():
		Found = False
		for cell in row:
			current_id = str(cell.value)
			if current_id in Index_Col:
				Key_Row = cell.row
				Found = True
				break

		if Found == True:
			break
	#print('Get compare column')		
	if Found == True:
		row = ws[Key_Row]
		ws.insert_rows(Key_Row+1)
		'''
		for temp_cell in ws[Key_Row+2]:
			merge_range = get_merge_range(ws, temp_cell)
			if merge_range != False:
				print('merge_range', merge_range)
				ws.unmerge_cells(merge_range)
		'''
		
		empty_col = 0
		for cell in row:
			if empty_col >50:
				print("No column more. Move to next step.")
				break
			current_cell_address = cell.coordinate
			#print('Current Cell', current_cell_address)
			main_key = None
			sub_key = None

			if type(ws[current_cell_address]).__name__ == 'MergedCell':
				main_key = getMergedCellVal(ws,  ws[current_cell_address])
			else:
				main_key = 	cell.value

			sub_cell_address = current_cell_address.replace(str(Key_Row), str(Key_Row-1))
			if type(ws[sub_cell_address]).__name__ == 'MergedCell':
				sub_key = getMergedCellVal(ws,  ws[sub_cell_address])
			else:
				sub_key = ws[sub_cell_address].value		
			#print('current_cell_address', current_cell_address)
			#print('sub_cell_address', sub_cell_address)
			#print('main_key', main_key)
			#print('sub_key', sub_key)
			new_cell_address = current_cell_address.replace(str(Key_Row), str(Key_Row+1))
			if sub_key in ['', None]:
				if main_key in ['', None]:
					empty_col+=1
					continue
				else:
					current_id = main_key.replace(' ', '_')
					current_id = main_key
					
			else:
				if sub_key in ['', None]:
					empty_col+=1
					continue
				else:
					main_key = main_key.replace(' ', '_')
					sub_key = sub_key.replace(' ', '_')
					
					if main_key == sub_key:
						current_id = main_key
					else:	
						current_id = sub_key + '.' + main_key 
	
			if current_id in List_Col:
				index = 1
				while True:
					temp_id = current_id + '_' + str(index)
					if temp_id not in List_Col:
						current_id = temp_id
						break
					index+=1
			empty_col = 0
			#print("Append col to list:", current_id)	
			List_Col.append(current_id)
			ws[new_cell_address].value = current_id
		#if len(List_Col) > 0:
			#print("Loading successful, break")
			#print('Remove row: ', Key_Row)
			#ws.delete_rows(Key_Row, 1)
		
	#print('Function_Fast_Create_Data', 'List_Col', List_Col)
	#print('Generate temp data file')
	temp = os.getcwd() + '/temp_' + Function_Get_TimeStamp() + '.xlsx'
	#print('temp:', temp)
	xlsx.save(temp)
	xlsx.close()
	#print('Key_Row', Key_Row)
	#print('Load data from temp file')
	#print('List_Col', List_Col)
	#print('Key_Row+1', Key_Row+1)
	#print()
	excel_data_df = pd.read_excel(temp, engine='openpyxl', sheet_name= Sheet_Name, skiprows = Key_Row, usecols = List_Col)

	#os.remove(temp)
	#print('excel_data_df', excel_data_df)
	#excel_data_df[main_index] =  excel_data_df[Index_Col].apply(lambda x: '.'.join(x.dropna().astype(str)),axis=1)	
	
	main_index = ''
	for index in Index_Col:
		if index == Index_Col[0]:
			main_index += index
		else:
			main_index += '.' + index

	if len(Index_Col) > 1:

		excel_data_df[main_index] = ''
		#print('Received data:', excel_data_df.columns.tolist())
		for index in Index_Col:
			#print('main_index', main_index)
			#print('index', index)
			arr_main_index = np.char.array(excel_data_df[main_index].values)
			arr_index = np.char.array(excel_data_df[index].values)
			
			if index == Index_Col[0]:
				excel_data_df[main_index] = (arr_main_index + arr_index).astype(str)
			else:
				excel_data_df[main_index] = (arr_main_index +  b'.' + arr_index).astype(str)
			
			excel_data_df = excel_data_df.drop(columns = [index], axis=1)
	
		#list_remove  = ['Id', '#Idx']
		#filtered_list_col = []
		#for index in list_remove:
		#	if index in excel_data_df.columns and index not in Index_Col:
		#		excel_data_df = excel_data_df.drop(columns = [index], axis=1)
			'''
			if index in List_Col:
				del List_Col[List_Col.index(index)]
			'''
	#print('main_index', main_index)
	#List_Col = [i for i in List_Col if i]
	List_Col = excel_data_df.columns.tolist()
	#del List_Col[List_Col.index(main_index)]
	if main_index in List_Col:
		List_Col.remove(main_index)
	List_Col.insert(0, main_index)

	#print('Received data:', List_Col)
	
	#print('List col: ',List_Col)
	# Remove temp file
	os.remove(temp)
	return excel_data_df, List_Col, main_index

def Function_Fast_Create_Data(Data_Workbook, Sheet_Name, Index_Col, Keep_Comment_Column = True):
	#print('Fast create data')
	#print('Source:', Data_Workbook)
	List_Col = []
	default_index = Index_Col[0]
	#print('default_index', default_index)
	#print('Function_Fast_Create_Data', 'Index_Col', Index_Col)
	print('Load workbook')
	xlsx = load_workbook(Data_Workbook, data_only = True)	
	#print('Load col list', Index_Col)
	ws = xlsx[Sheet_Name]
	for row in ws.iter_rows():
		for cell in row:
			current_id = str(cell.value)
			#print('current_id', current_id)
			if current_id in Index_Col:
				
				Key_Row = cell.row
				empty_col_count = 0
				for cell in row:
					current_id = str(cell.value)
					print('current_id:', cell.coordinate, current_id)
					if current_id != 'None':
						if Key_Row > 0:
							CellAddress = cell.column_letter + str(cell.row - 1)
							
							try:

								if type(ws[CellAddress]).__name__ == 'MergedCell':
									#print('Merged cell')
									try:
										main_group = getMergedCellVal(ws,  ws[CellAddress])
									except Exception as e:
										print('Error: ', e)
								else:
									#print('Normal cell')
									# By checking if the cell is merged or not, we can ignore the comment above the column name.
									#main_group = ws[CellAddress].value
									main_group = None
								#print('main_group', main_group)
							except Exception as E:
								print('Error:', E)
								main_group = None
							
							if main_group != None:
								if current_id in Index_Col:
									Index_Col[Index_Col.index(current_id)] = main_group + '.' + current_id
								current_id = main_group + '.' + current_id
								cell.value = current_id	
							if Keep_Comment_Column == True:
								if current_id.startswith("#"):
									if current_id in Index_Col:
										List_Col.append(current_id)
								else:
									List_Col.append(current_id)	
							else:
								List_Col.append(current_id)	
							#print('current_id',current_id)
					else:
						empty_col_count +=1
						if empty_col_count >= 50:
							print('Too much empty col, skip.')
							break
			if len(List_Col) > 0:
				print("Loading successful, break")
				break
		if len(List_Col) > 0:
				print("Loading successful, break")
				break		
	#print('Function_Fast_Create_Data', 'List_Col', List_Col)
	print('Generate temp data file')
	temp = os.getcwd() + '/temp_' + Function_Get_TimeStamp() + '.xlsx'
	print('temp:', temp)
	xlsx.save(temp)
	xlsx.close()
	#print('Key_Row', Key_Row)
	print('Load data from temp file')
	print('List_Col', List_Col)
	excel_data_df = pd.read_excel(temp, engine='openpyxl', sheet_name= Sheet_Name, skiprows = Key_Row-1, usecols = List_Col)
	#os.remove(temp)
	print('excel_data_df', excel_data_df)
	#excel_data_df[main_index] =  excel_data_df[Index_Col].apply(lambda x: '.'.join(x.dropna().astype(str)),axis=1)	
	
	main_index = ''
	for index in Index_Col:
		if index == Index_Col[0]:
			main_index += index
		else:
			main_index += '.' + index

	if len(Index_Col) > 1:

		excel_data_df[main_index] = ''
		#print('Received data:', excel_data_df.columns.tolist())
		for index in Index_Col:
			#print('main_index', main_index)
			#print('index', index)
			arr_main_index = np.char.array(excel_data_df[main_index].values)
			arr_index = np.char.array(excel_data_df[index].values)
			
			if index == Index_Col[0]:
				excel_data_df[main_index] = (arr_main_index + arr_index).astype(str)
			else:
				excel_data_df[main_index] = (arr_main_index +  b'.' + arr_index).astype(str)
			
			excel_data_df = excel_data_df.drop(columns = [index], axis=1)
	
		list_remove  = ['Id', '#Idx']
		#filtered_list_col = []
		for index in list_remove:
			if index in excel_data_df.columns and index not in Index_Col:
				excel_data_df = excel_data_df.drop(columns = [index], axis=1)
			'''
			if index in List_Col:
				del List_Col[List_Col.index(index)]
			'''
	#print('main_index', main_index)
	#List_Col = [i for i in List_Col if i]
	List_Col = excel_data_df.columns.tolist()
	#del List_Col[List_Col.index(main_index)]
	if main_index in List_Col:
		List_Col.remove(main_index)
	List_Col.insert(0, main_index)

	print('Received data:', List_Col)
	
	#print('List col: ',List_Col)

	return excel_data_df, List_Col, main_index

def getMergedCellVal(sheet, cell):
	rng = [s for s in sheet.merged_cells.ranges if cell.coordinate in s]
	if len(rng)!=0:
		current_cell =  sheet.cell(rng[0].min_row, rng[0].min_col)
		color_in_hex = current_cell.fill.start_color.index # this gives you Hexadecimal value of the color
		if color_in_hex != '00000000':
			return current_cell.value
		else:
			return None	
	else:
		color_in_hex = cell.fill.start_color.index # this gives you Hexadecimal value of the color
		print('Background colour: ', color_in_hex)
		if color_in_hex != '00000000':
			return cell.value
		else:
			return None		

def get_merge_range(sheet, cell):
	rng = [s for s in sheet.merged_cells.ranges if cell.coordinate in s]
	if len(rng)!=0:
		start_add = get_column_letter(rng[0].min_col) + str(rng[0].min_row)
		end_add = get_column_letter(rng[-1].max_col) + str(rng[-1].max_row)
		merge_range = start_add + ':' + end_add
		return merge_range
	return False

def assign_value(sheet, cell):
	value = getMergedCellVal(sheet, cell)
	if value != None:
		rng = [s for s in sheet.merged_cells.ranges if cell.coordinate in s]
		if len(rng)!=0:
			for row_index in range(rng[0].min_row, rng[-1].max_row):
				for col_index in range(rng[0].min_col, rng[-1].max_col):
					cell_add =  get_column_letter(col_index) + str(row_index)
					sheet[cell_add].value = value
	return sheet		

def Add_Sheet():
	return

def Function_Get_TimeStamp():		
	now = datetime.now()
	timestamp = str(int(datetime.timestamp(now)))			
	return timestamp

def Function_Add_Surflix(File, SubFolder = None, Surflix = None, add_time_stamp = True):
	
	Outputdir = os.path.dirname(File)
	baseName = os.path.basename(File)
	sourcename, ext = os.path.splitext(baseName)

	if SubFolder != None:
		Outputdir += '/' + SubFolder
	
	if not os.path.isdir(Outputdir):
		try:
			os.mkdir(Outputdir)
		except OSError:
			Outputdir = Outputdir
	if Surflix != None:
		sourcename +=  "_" + Surflix
	if add_time_stamp:
		timestamp = Function_Get_TimeStamp()
		sourcename += "_" + str(timestamp)
	Name = Outputdir + '/' + sourcename + ext

	return Name

###########################################################################################
# AUTO TEST
###########################################################################################

def Function_AutoTest( Status_Queue, Process_Queue, test_configuration_file, old_data_folder, new_data_folder, Background_Color, Font_Color):
	Status_Queue.put('Load test configuration.')
	test_configuration = load_test_configuration(test_configuration_file, old_data_folder, new_data_folder)
	Status_Queue.put('Configuration done')
	Status_Queue.put('Execute test: ')
	Start = time.time()
	
	Sum_Result = {}
	now = datetime.now()
	testcase_name = os.path.basename(test_configuration_file)
	_raw_testcase_name, raw_ext = os.path.splitext(testcase_name)

	output_folder = new_data_folder + '\\' + _raw_testcase_name + '_' + str(int(datetime.timestamp(now)))
	os.mkdir(output_folder)
	total_task = len(test_configuration)
	current_task = 0

	for index_number in range(len(test_configuration)):
		testcase = test_configuration[index_number]
		Status_Queue.put('Testcase name: ' + str(testcase))

		base_name = testcase['file']
		test_type = testcase['function']
		Sheet_Name = testcase['data_sheet']
		Index_Col = testcase['key']
		Index_Col_List = Index_Col.split(',')
		
		Sum_Result[base_name] = {}
		Sum_Result[base_name]['Type'] = test_type
		Sum_Result[base_name]['Result'] = {}

		if test_type == 'Compare':
			file_name = testcase['file']
			Old_File = old_data_folder + '\\' + file_name
			New_File = new_data_folder + '\\' + file_name
			if os.path.isfile(Old_File):
				if not os.path.isfile(New_File):
					Sum_Result[base_name]['Result']['Changed'] = True
					Sum_Result[base_name]['Result']['Type'] = 'Dropped'
					continue
			else:
				if not os.path.isfile(New_File):
					Sum_Result[base_name]['Result']['Changed'] = True
					Sum_Result[base_name]['Result']['Type'] = 'FileNotFound(All)'
					continue
				else:
					Sum_Result[base_name]['Result']['Changed'] = True
					Sum_Result[base_name]['Result']['Type'] = 'Added'
					continue
			try:
				index_list_str = '_'.join(Index_Col_List)
				raw_file_name, ext_name = os.path.splitext(file_name)
				output_result = output_folder + '\\' + raw_file_name + '_' + index_list_str + ext_name
				test_result = Function_Fast_Compare_Single_Data(Status_Queue, Old_File, New_File, Sheet_Name, Index_Col_List, output_result, True, Background_Color, Font_Color)

				Sum_Result[base_name]['Result'] = test_result
			except Exception as e:
				print("error:", e)
				Sum_Result[base_name]['Result']['Changed'] = True
				Sum_Result[base_name]['Result']['Type'] = "ERROR"
				test_result = "Fail"
			
		
		elif test_type == 'Language Validate':

			File = old_data_folder + '\\' + testcase['file']
			try:
				test_result = Function_Validate_Language(File, testcase['key'])
			except:
				Sum_Result[base_name]['Result']['Changed'] = True
				Sum_Result[base_name]['Result']['Type'] = "ERROR"
				continue

			if test_result > 0:
				Sum_Result[base_name]['Result']['Changed'] = True
				Sum_Result[base_name]['Result']['Details'] = test_result
			else:
				Sum_Result[base_name]['Result']['Changed'] = False
		else:
			continue	
		
		Status_Queue.put('Testcase result: ' + str(test_result))	
		current_task += 1
		update_progress(Process_Queue, current_task, total_task)

	summary = Workbook()
	ws =  summary.active
	ws.title = 'Summary'
	Header = ['File', 'Type', 'Result', 'Changed', 'Added', 'Dropped']
	Row = 1
	Col = 2
	ws.cell(row=Row, column=Col).value = "Old data: " + str(old_data_folder)
	Row += 1
	ws.cell(row=Row, column=Col).value = "New data: " + str(new_data_folder)

	Row = 3

	for Par in Header:
		ws.cell(row=Row, column=Col).value = Par
		Col +=1
	ws.column_dimensions['B'].width = 50
	ws.column_dimensions['C'].width = 20
	ws.column_dimensions['D'].width = 15

	Row +=1
	column_letters = ['B', 'C', 'D', 'E', 'F', 'G']

	for column_letter in column_letters:
		ws.column_dimensions[column_letter].bestFit = True

	for file_name in Sum_Result:
		ws.cell(row=Row, column=2).value = file_name
		
		Type = Sum_Result[file_name]['Type']	
		ws.cell(row=Row, column=3).value = Type

		Changed = Sum_Result[file_name]['Result']['Type']
		ws.cell(row=Row, column=4).value = Changed
		
		if Changed:
			if 'Details' in Sum_Result[file_name]['Result']:
				if Type == 'Compare':
					Details = Sum_Result[file_name]['Result']['Details']
					ws.cell(row=Row, column=5).value = Details['Diff']
					ws.cell(row=Row, column=6).value = Details['Added']
					ws.cell(row=Row, column=7).value = Details['Dropped']
				# Language validate result
				else:
					ws.cell(row=Row, column=5).value = Changed
		Row +=1
	
	Tab = Table(displayName="Summary", ref="B2:" + "G" + str(Row-1))
	style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=True)
	Tab.tableStyleInfo = style
	ws.add_table(Tab)

	summary.save(output_folder + '\\Summary.xlsx')	
	summary.close()
	Status_Queue.put("Result found at: " + str(output_folder + '\\Summary.xlsx'))
	End = time.time()
	Total = int(End - Start)
	Status_Queue.put("'Time spent: " + str(Total))	


def update_progress(Progress_Queue, current_progress, total_progress):
	print('Total:', total_progress)
	print('Current:', current_progress)
	percent = int(1000 * current_progress / total_progress)
	print('Current percent: ', percent)
	Progress_Queue.put(percent)

def load_test_configuration(test_configuration_file, old_folder, new_folder):
	test_configuration = []
	xlsx = load_workbook(test_configuration_file, data_only=True)
	type_name_column = None
	type_name_row = None
	try:
		testcase_sheet = xlsx['Testcase']
	except:
		return []
	for row in testcase_sheet.iter_rows():
		for cell in row:
			value = str(cell.value)
			if value == 'Testcase':
				type_name_row = cell.row
				type_name_column = cell.column_letter
				break
			else:
				continue
	if type_name_column == None:
		return []
	else:
		total_row = testcase_sheet.max_row

		idx = column_index_from_string(type_name_column)
		for row_index in range(type_name_row+1, total_row+1):

			testcase = {}
			function_name = testcase_sheet[get_column_letter(idx) + str(row_index)].value
	
			if function_name == None:
				continue
			testcase['function'] = testcase_sheet[get_column_letter(idx) + str(row_index)].value
			file_name = testcase_sheet[get_column_letter(idx+1) + str(row_index)].value

			if file_name == None:
				continue
			testcase['file'] = file_name
			data_sheet = testcase_sheet[get_column_letter(idx+2) + str(row_index)].value

			if data_sheet == None:
				data_sheet = 'Data'
			testcase['data_sheet'] = data_sheet
			key = testcase_sheet[get_column_letter(idx+3) + str(row_index)].value
			if key == None:
				key = 'StringId'
			testcase['key'] = key

			test_configuration.append(testcase)

	return test_configuration

###########################################################################################
# DEEP COMPARE DATA
###########################################################################################
def Function_Create_Data(Data_Workbook, Sheet_Name, Index_Col,):

	Data_Dict = {}
	xlsx = Data_Workbook
	Col_List = {}
	Col_Count = {}
	Param_List = []
	ID_Col = None
	ID_Row = None
	database = None
	Start = time.time()
	for sheet in xlsx:
		sheetname = sheet.title
	
		if sheetname == Sheet_Name:

			ws = xlsx[sheet.title]

			for row in ws.iter_rows():
				for cell in row:
					if cell.value == Index_Col:
						ID_Col = cell.column_letter
						ID_Row = cell.row
						#print('ID row:',ID_Row)
						database = ws
						break

				if database!=  None:
					break

			if ID_Col == None:
				return None

			for cell in ws[ID_Row]:
				if cell.value not in ['', Index_Col]:
					Col_Val = cell.value
					if Col_Val == None:
						Col_Val = 'None'
					try:
						Count = Col_Count[Col_Val]
					except:
						Col_Count[Col_Val] = 1
						Count = 1
					
	
					if Col_Val not in Param_List:
						Col_List[cell.column_letter] = Col_Val
						Param_List.append(Col_Val)
					else:
						Col_Count[Col_Val] +=1

						Col_List[cell.column_letter] = str(Col_Val) + '[' + str(Count) + ']'


	Key_List = Col_List.keys()
	#print(Col_List)
	if database == None:
		return {}
	else:
		EmptyRow = 0	
		Data_Frame = {}
		Data_Sheet = xlsx[Sheet_Name]
		MaxRow = Data_Sheet.max_row
		for Row in range(ID_Row+1, MaxRow+1):
			if EmptyRow <= 20:
				Row_Data =  Data_Sheet[Row]
				Current_ID =  Data_Sheet[ID_Col+str(Row)].value
				if Current_ID != None:
					#print('Adding ID:', Current_ID)
					Entry = {}
					#Entry['ID'] = Current_ID
					for Pair in Key_List:
						Value = Data_Sheet[Pair+str(Row)].value
						if Value in ["", None]:
							Value = '#N/A'
						Entry[Col_List[Pair]] = Value

					Data_Frame[Current_ID] = Entry
					EmptyRow = 0
				else:
					EmptyRow+=1
			else:
				break		
		'''
		
		'''
		End = time.time()
		Total = End - Start
		#print('Time spent:', Total)
		#print('Col_List', Col_List)
		#print('Data_Frame', Data_Frame)
		return 	{'Label': Col_List, 'Data':  Data_Frame}

	
	'''

			if database != None:
				for i in range(KR_Row, database.max_row): 
					KRAddress = KR_Coll + str(i+1)
					ENAddress = EN_Coll + str(i+1)
					print('KRAddress', KRAddress)
					print('ENAddress', ENAddress)
					KRCell = database[KRAddress]
					KRValue = KRCell.value
					ENCell = database[ENAddress]
					ENValue = ENCell.value
					if KRValue == None or ENValue == None or KRValue == 'KO' or ENValue == 'EN':
						continue
					elif KRValue != None and ENValue != None:
						Dict.append([KRValue, ENValue.lower()])
	print("Successfully load dictionary from: ", DictList)
	return Dict
		
	'''

def Deep_Compare_Data(Old_Data, New_Data):

	Old_Set = Old_Data['Data']
	New_Set = New_Data['Data']

	Removed = []
	Removed_Data =  {}
	Added = []
	Added_Data =  {}
	ToCheck = []
	Changed = []
	Changed_Data = {}
	Old_List = list(Old_Set.keys())

	New_List = list(New_Set.keys())

	for ID in Old_List:
		if ID not in New_List:
			#print('Add', ID, 'to Removed')
			Removed_Data[ID] = Old_Set[ID]
		else:
			#print('Add', ID, 'to toCheck')
			ToCheck.append(ID)
	

	for ID in New_List:
		if ID not in Old_List:
			#print('Add', ID, 'to Added')
			Added_Data[ID] = New_Set[ID]
	
	for ID in ToCheck:
		ChangeFlag = False
		Old_ID_Data = Old_Set[ID]

		New_ID_Data = New_Set[ID]
		
		Old_Param_List = list(Old_ID_Data.keys())

		New_Param_List = list(New_ID_Data.keys())
	
		for Par in Old_Param_List:
			#print(Par)
			if Par not in New_Param_List:
				New_ID_Data[Par] = 'Removed value --> ' + str(Old_ID_Data[Par])
				ChangeFlag = True
				
			else:
				if New_ID_Data[Par] != Old_ID_Data[Par]:
					#print('Old: ', str(Old_ID_Data[Par]))
					#print('New: ', str(New_ID_Data[Par]))
					New_ID_Data[Par] = str(Old_ID_Data[Par]) + ' --> ' + str(New_ID_Data[Par])
					ChangeFlag = True
		
		for Par in New_Param_List:
			#print(Par)
			if Par not in Old_Param_List:
				#Added.append(ID)
				New_ID_Data[Par] = 'New added --> ' + str(New_ID_Data[Par])
				ChangeFlag = True

		if ChangeFlag:
			Changed_Data[ID] = New_Set[ID]
			

	#for ID in Changed:
	#	Changed_Data[ID] = New_Set[ID]
		#Changed_Data.append({'ID':New_Set[ID]})
	#print(Changed_Data)
	Diff = {
		'Removed': Removed_Data,
		'Added': Added_Data,
		'Changed': Changed_Data
	}
	return Diff

def Function_Print_Excel(Data_Frame):

	for x in Data_Frame:
		print(x)
		print(Data_Frame[x])
		print('\n')
		#for ID in Data_Frame[x]:
		#	print(ID)
		#	print('\n')

	return


def Function_Deep_Compare_Data(
		Status_Queue, Process_Queue, Old_Files, New_Files, ExceptionList, out_path, Sheet_Name, Index_Col_Name, Background_Colour, Font_Colour, **kwargs):
	Start = time.time()

	my_color = Color(rgb=Background_Colour)
	my_fill = PatternFill(patternType='solid', fgColor=my_color)

	my_font = Font(color=Font_Colour)
	print('ExceptionList', ExceptionList)
	for Old_File in Old_Files:
		if Old_File != None:
			print('Old_File', Old_File)
			if Old_File in ExceptionList:
				continue
			Old_Outputdir = os.path.dirname(Old_File)
			Old_baseName = os.path.basename(Old_File)
			Old_sourcename, Old_ext = os.path.splitext(Old_baseName)

			for New_File in New_Files: 
				if New_File != None:
					print('Checking: ', New_File)
					#New_Outputdir = os.path.dirname(New_File)
					
					New_baseName = os.path.basename(New_File)
					
					#New_sourcename, New_ext = os.path.splitext(New_baseName)
					if Old_baseName == New_baseName:
						Status_Queue.put("Loading old data: " + New_baseName)
						Old_DB = load_workbook(Old_File, data_only=True)

						Status_Queue.put("Loading new data: " + New_baseName)
						New_DB = load_workbook(New_File, data_only=True)

						Status_Queue.put("Import new data: " + New_baseName)
						New_Data = Function_Create_Data(New_DB, Sheet_Name, Index_Col_Name)
						if New_Data == None:
							Status_Queue.put("Invalid New Data's structure.")
						#Function_Print_Excel(New_Data)
						
						Status_Queue.put("Import old data" + New_baseName)
						Old_Data = Function_Create_Data(Old_DB, Sheet_Name, Index_Col_Name)
						if Old_Data == None:
							Status_Queue.put("Invalid Old Data's structure.")

						#Function_Print_Excel(Old_Data)
						Status_Queue.put("Compare data for " + New_baseName)
						Diff_Data = Deep_Compare_Data(Old_Data, New_Data)
						Status_Queue.put("Generate compare result for: " + New_baseName)

						Changed = Diff_Data['Changed']
						Removed = Diff_Data['Removed']
						Added = Diff_Data['Added']

						wb = Workbook()
						ws =  wb.active
						ws.title = 'Changed'
						Row = 2
						Changed_List = list(Changed.keys())
						Header = []
						for Sample in Changed:
							Header = list(Changed[Sample].keys())
							break
						Header = [Index_Col_Name] + Header
						Row = 2
						if Changed != {}:
							ws.sheet_properties.tabColor = Background_Colour
							Col = 1
							for Par in Header:
								ws.cell(row=1, column=Col).value = Par
								Col +=1


							for item in Changed:
								ws.cell(row=Row, column=1).value = item
								Col = 2	
								for Par in Changed[item]:
									Value = Changed[item][Par]
									ws.cell(row=Row, column=Col).value = Value = Changed[item][Par]
									if '-->' in str(Value) :
										ws.cell(row=Row, column=Col).fill = my_fill
										ws.cell(row=Row, column=Col).font = my_font
										#Font_Colour
									LastCell = ws.cell(row=Row, column=Col).column_letter
									Col+=1
								Row +=1

							Tab = Table(displayName="Change", ref="A1:" + LastCell + str(Row-1))
							style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=True)
							Tab.tableStyleInfo = style
							ws.add_table(Tab)

						
						
						if Added != {}:

							wb.create_sheet('Added')
							ws = wb['Added']
							#Added_List = list(Added.keys())
							Header = []
							for Sample in Added:
								Header = list(Added[Sample].keys())
								break
							Header = [Index_Col_Name] + Header
							Row = 2

							ws.sheet_properties.tabColor = Background_Colour
							Col = 1
							for Par in Header:
								ws.cell(row=1, column=Col).value = Par
								Col +=1


							for item in Added:
								ws.cell(row=Row, column=1).value = item
								Col = 2	
								for Par in Added[item]:
									ws.cell(row=Row, column=Col).value = Added[item][Par]
									Col+=1
									LastCell = ws.cell(row=Row, column=Col).column_letter
								Row +=1
						
							Tab = Table(displayName="Added", ref="A1:" + LastCell + str(Row-1))
							style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=True)
							Tab.tableStyleInfo = style
							ws.add_table(Tab)

						if Removed != {}:
							wb.create_sheet('Removed')
							#Removed_List = list(Removed.keys())
							ws = wb['Removed']
							Header = []
							for Sample in Removed:
								Header = list(Removed[Sample].keys())
								break
							Header = [Index_Col_Name] + Header
							Row = 2
							ws.sheet_properties.tabColor = Background_Colour
							Col = 1
							for Par in Header:
								ws.cell(row=1, column=Col).value = Par
								Col +=1


							for item in Removed:
								ws.cell(row=Row, column=1).value = item
								Col = 2	
								for Par in Removed[item]:
									ws.cell(row=Row, column=Col).value = Removed[item][Par]
									Col+=1
									LastCell = ws.cell(row=Row, column=Col).column_letter
								Row +=1
							
							Tab = Table(displayName="Removed", ref="A1:" + LastCell + str(Row-1))
							style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=True)
							Tab.tableStyleInfo = style
							ws.add_table(Tab)


						NewName = Function_Add_Surflix(New_File, 'Compare Result', "Deep_Compare")
						if Removed == {} and Added == {} and Changed == {}:
							Status_Queue.put("No change.")
						else:
							try:			
								wb.save(NewName)	
								Status_Queue.put("Done!")

							except:
								Status_Queue.put("Permission denied, fail to save result file!")
	End = time.time()
	Total = End - Start
	Status_Queue.put("Total time spend: " + str(Total))


###########################################################################################
# BAD WORD FUNCTION
###########################################################################################
def BadWord_Load_BadWord_DB(DB_Path,):

	if DB_Path != None:
		if (os.path.isfile(DB_Path)):
			
			xlsx = load_workbook(DB_Path)	
			Exception = []
			for sheet in xlsx:
				print("Loading DB from:", sheet.title)
				database = None
				TextColl = ""
				ws = xlsx[sheet.title]
				for row in ws.iter_rows():
					
					for cell in row:
						text = str(cell.value).lower()
						if text == "string":
							TextColl = cell.column_letter
							Row = cell.row
							database = ws
							break	
					if database != None:
						break	

				if database != None:
					for i in range(Row, database.max_row): 
						ExceptionAdderss = TextColl + str(i+1)
						ExceptionCell = database[ExceptionAdderss]
						ExceptionValue = ExceptionCell.value
						if ExceptionValue == None:
							continue
						else:
							Exception.append(ExceptionValue)
			return Exception
			
		else:
			print('Exception list is not existed')
			return []
	else:
		return []

'''
def Function_BadWord_Execute_MultiThread(
		Status_Queue, Process_Queue, Text_Files, DB_File, output_folder, Sheet_Name, Index_Col_Name, Background_Colour, Font_Colour, **kwargs):

	Index_Col_List = Index_Col_Name.split(',')
	Data_Sheet = Sheet_Name
	my_yellow_color = Color(rgb= 'ffff00')
	my_yellow_fill = PatternFill(patternType='solid', fgColor=my_yellow_color)

	_db_list = BadWord_Load_BadWord_DB(DB_File)
	number_of_processes = multiprocessing.cpu_count()

	while len(TaskList) > 0:
		if len(TaskList) > number_of_processes:
			NewTask = number_of_processes
		else:
			NewTask = len(TaskList)

		for w in range(NewTask):
			input_file = TaskList[0]

			baseName = os.path.basename(input_file)
			sourcename, ext = os.path.splitext(baseName)
			Status_Queue = output_folder + '//' + sourcename + ext
			StatusQueue.put('Process file: ' + sourcename)	
			p = Process(target= BadWord_Check_Single_File, args=(input_file, Index_Col_List, _db_list, result_file,))
			del TaskList[0]
			processes.append(p)
			p.start()

		for p in processes:
			p.join()
			Complete+=1
		
		percent = ShowProgress(Complete, TotalFile)
		Process_Queue.put(percent)
		
	Status_Queue.put('Optimized done.')
'''

def Function_BadWord_Execute(
		Status_Queue, Process_Queue, Text_Files, DB_File, output_folder, Sheet_Name, Index_Col_Name, exact_match, Background_Colour, Font_Colour, **kwargs):

	Index_Col_List = Index_Col_Name.split(',')
	Data_Sheet = Sheet_Name
	my_yellow_color = Color(rgb= 'ffff00')
	my_yellow_fill = PatternFill(patternType='solid', fgColor=my_yellow_color)

	_db_list = BadWord_Load_BadWord_DB(DB_File)
	print("length of DB:", len(_db_list))
	TotalFile = len(Text_Files)
	Sum_Result = {}
	Complete = 0

	for FileName in Text_Files:
		Source_Name = os.path.basename(FileName)

		Sum_Result[Source_Name] = []

		baseName = os.path.basename(FileName)
		sourcename, ext = os.path.splitext(baseName)
		result_file = output_folder + '/' + sourcename + ext
		Status_Queue.put('Checking: ' + sourcename )
		Compare_Result = BadWord_Check_Single_File(FileName, Data_Sheet, Index_Col_List, exact_match,_db_list, result_file, my_yellow_fill, my_yellow_color,)

		Sum_Result[Source_Name] = Compare_Result
		Complete+=1
		percent = ShowProgress(Complete, TotalFile)
		Process_Queue.put(percent)
		#Status_Queue.put('Optimized done.')

	summary = Workbook()
	ws =  summary.active
	ws.title = 'Summary'
	Header = ['File', 'Result', 'Count', 'Bad Words found']
	Col = 2
	Row = 2
	for Par in Header:
		ws.cell(row=Row, column=Col).value = Par
		Col +=1
	Row +=1
	column_letters = ['B', 'C', 'D', 'E']

	if exact_match:
		ws.cell(row=1, column=2).value = "Match type: Exact match"
	else:
		ws.cell(row=1, column=2).value = "Match type: Text contains"
	
	for file_name in Sum_Result:
	
		ws.cell(row=Row, column=2).value = file_name
		Result = "Pass"
		if len(Sum_Result[file_name]) > 0:
			Result = "Fail"
		
		ws.cell(row=Row, column=3).value = Result

		if Result == "Fail":
			ws.cell(row=Row, column=4).value = len(Sum_Result[file_name])
			Sum_Result[file_name] = list(dict.fromkeys(Sum_Result[file_name]))
			_list_bad_Words = ",".join(Sum_Result[file_name])
			ws.cell(row=Row, column=5).value = _list_bad_Words
				
		Row +=1
	
	Tab = Table(displayName="Summary", ref="B2:" + "E" + str(Row-1))
	style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=True)
	Tab.tableStyleInfo = style
	ws.add_table(Tab)

	for column_letter in column_letters:
		ws.column_dimensions[column_letter].bestFit = True

	now = datetime.now()
	timestamp = str(int(datetime.timestamp(now)))	


	summary_result = output_folder + '/' + 'Bad word check result_'+ timestamp + '.xlsx'
	summary.save(summary_result)	
	summary.close()

	Status_Queue.put("Result found at: " + str(summary_result))

def BadWord_Check_Single_File(text_file_path, Data_Sheet_Name, column_id, exact_match, db_list, result_file, Background_Colour, Font_Colour, **kwargs ):

	xlsx = load_workbook(text_file_path, data_only=True)
	print("Exact match: ", exact_match)
	ID = None
	_column_list = []
	First_Row = None

	Data_Sheet = xlsx[Data_Sheet_Name]
	for row in Data_Sheet.iter_rows():
		for cell in row:
			value = str(cell.value)
			if value in column_id:
				First_Row = cell.row
				_column_list.append(cell.column_letter)
				_result_col = int(cell.column)+1
				Data_Sheet.insert_cols(_result_col)
				Data_Sheet[get_column_letter(_result_col) + str(First_Row)].value = value + '_Result'
			else:
				continue
	
	Last_Row = Data_Sheet.max_row

	if ID == None:
		ID= _column_list[0]
	bad_word_check_result = []
	if First_Row != None:

		ShowRow = []
		
		for row in Data_Sheet[ID + str(First_Row) +':' + ID + str(Last_Row)]:

			for cell in row:
				value = cell.value
				row_index = cell.row

				for check_col in _column_list:
					String_Detail = Data_Sheet[check_col+str(row_index)].value
					if String_Detail != None:
						temp_result = Bad_Word_Check(String_Detail, db_list, exact_match)
					
						if len(temp_result) > 0:
							
							_result_col =get_column_letter(Data_Sheet[check_col+str(row_index)].column + 1)
							Data_Sheet[_result_col+str(row_index)].value = ""
							Data_Sheet[_result_col+str(row_index)].value += '[Bad words found]: \n'	
							Data_Sheet[_result_col+str(row_index)].value += ','.join(temp_result)
							Data_Sheet[_result_col+str(row_index)].fill = Background_Colour
							
							if isinstance(Data_Sheet.row_dimensions[row_index].height, int):
								Data_Sheet.row_dimensions[row_index].height += 10
							ShowRow.append(row_index)
							bad_word_check_result += temp_result
						else:
							Data_Sheet.row_dimensions[row_index].hidden= True	
					else:
						Data_Sheet.row_dimensions[row_index].hidden= True
	
		if len(ShowRow) > 0:	
			xlsx.save(result_file)

	return bad_word_check_result

def Bad_Word_Check(text, db_list, exact_math = False):
	'''Check if a text is in the bad word list or not, return the list of bad word show'''

	if not isinstance(text, str):
		return []
	bad_word_found = []
	for word in db_list:
		if exact_math:
			if word == text:
				bad_word_found.append(word)
		else:
			if word in text:
				bad_word_found.append(word)
	return bad_word_found		

###########################################################################################
def Function_Validate_Language(Data_Workbook, Key_ID, Value_List = None, Column_ID = None):

	print('Data_Workbook', Data_Workbook)
	print('Key_ID', Key_ID)
	print('Value_List', Value_List)
	print('Column_ID', Column_ID)


	my_yellow_color = Color(rgb= 'ffff00')
	my_yellow_fill = PatternFill(patternType='solid', fgColor=my_yellow_color)

	xlsx = load_workbook(Data_Workbook, data_only=True)
	#print(Show_ID, Column_ID)
	ID = None
	Column_List = []
	First_Row = None

	Data_Sheet = xlsx['Data']
	for row in Data_Sheet.iter_rows():
		for cell in row:
			value = str(cell.value)
			if value in Column_ID:
				First_Row = cell.row
				Column_List.append(cell.column_letter)
			elif value == Key_ID:
				ID = cell.column_letter
			else:
				continue
	Last_Row = Data_Sheet.max_row
	if ID == None:
		ID = "A"
	if First_Row != None:
		
		ShowRow = []
		for Col in Data_Sheet[ID + str(First_Row) +':' + ID + str(Last_Row)]:

			for cell in Col:
				value = cell.value

				if In_string(value, Value_List) or len(Value_List) == 0:
					row = cell.row
					RowPass = True
					for check_col in Column_List:
						String_Detail = Data_Sheet[check_col+str(row)].value
						if ValidateKoreanSource(String_Detail):
							Data_Sheet[check_col+str(row)].fill = my_yellow_fill
							RowPass = False
					if not RowPass:
						ShowRow.append(row)

		for row in Data_Sheet.iter_rows():
			for cell in row:
				row = cell.row
				if row != First_Row+1:
					break
				try:
					col_letter = cell.column_letter
				except:
					continue
				if col_letter == ID or col_letter in Column_List:
					
					continue
				else:
					Data_Sheet.column_dimensions[cell.column_letter].hidden= True
					continue
		for row_num in range(First_Row+1, Last_Row+1):
			if row_num not in ShowRow:
				Data_Sheet.row_dimensions[row_num].hidden= True
		if len(ShowRow) >0:
			timestamp = Function_Get_TimeStamp()	
			path = os.path.dirname(Data_Workbook) + "//Validation Result"
			
			if not os.path.isdir(path):
				try:
					os.mkdir(path)
				except OSError:
					print ("Creation of the directory %s failed" % path)
			else:
				print('Roaming folder exist.')

			baseName = os.path.basename(Data_Workbook)
			filename, ext = os.path.splitext(baseName)
			Output_Result = path + '/' + filename + '_' + str(timestamp) + '.xlsx'
			xlsx.save(Output_Result)
			return len(ShowRow)
		else:
			print('No change')	
			return 0

def In_string(search_string, string_list):
	for each_str in string_list:
		if each_str in search_string:
			return True
	return False		


def ValidateKoreanSource(string):
	if string not in [None, ""]:
		for i in range(len(string)):
			c = string[i]
			if unicodedata.category(c)[0:2] == 'Lo' : # other characters
				try:
					if 'HANGUL' in unicodedata.name(c) : return True
				except:
					continue
		return False

def ValidateEnglishSource(string):
	for i in range(len(string)):
		c = string[i]
		if unicodedata.category(c)[0:2] == 'Ll' : # other characters
			try:
				if 'LATIN' in unicodedata.name(c) : return True
			except:
				continue
	return False

###########################################################################################

def main():
	Process_Queue = Queue()
	Result_Queue = Queue()
	Status_Queue = Queue()
	Debug_Queue = Queue()
	
	MyManager = Manager()
	Default_Manager = MyManager.list()
	
	root = Tk()
	My_Queue = {}
	My_Queue['Process_Queue'] = Process_Queue
	My_Queue['Result_Queue'] = Result_Queue
	My_Queue['Status_Queue'] = Status_Queue
	My_Queue['Debug_Queue'] = Debug_Queue

	My_Manager = {}
	My_Manager['Default_Manager'] = Default_Manager

	Document_Utility(root, Queue = My_Queue, Manager = My_Manager,)
	root.mainloop()  


if __name__ == '__main__':
	if sys.platform.startswith('win'):
		multiprocessing.freeze_support()

	main()
